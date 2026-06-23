"""
Personal Finance Dashboard — finance.yeahia.uk
Flask app on port 8082
"""
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
import sqlite3, json, os, datetime, io, re, random
from pathlib import Path
from werkzeug.security import generate_password_hash, check_password_hash

BASE = Path(__file__).parent
DATA = BASE / "data"
DATA.mkdir(exist_ok=True)
DB   = DATA / "finance.db"
SETTINGS = DATA / "providers.json"
TAX_NOTES_FILE = DATA / "tax_notes.json"

# Live FX rates — fetched from open.er-api.com (free, no key) and cached for 1 hour.
# Fallback to manual snapshots if the API is unreachable.
import json as _json, time as _time, threading as _threading
_FX_CACHE_FILE = DATA / "fx_cache.json"
_FX_LOCK = _threading.Lock()
_FX_FALLBACK = {"USD_AUD": 1.4275, "BDT_AUD": 1 / 86.15, "ts": 0}

def _load_fx_cache():
    try:
        if _FX_CACHE_FILE.exists():
            return _json.loads(_FX_CACHE_FILE.read_text())
    except Exception:
        pass
    return dict(_FX_FALLBACK)

def _save_fx_cache(data):
    try:
        _FX_CACHE_FILE.write_text(_json.dumps(data))
    except Exception:
        pass

def _fetch_live_fx():
    """Fetch live rates from open.er-api.com in a background thread. Caches for 1 hour."""
    import urllib.request
    try:
        url = "https://open.er-api.com/v6/latest/USD"
        req = urllib.request.Request(url, headers={"User-Agent": "FinanceHub/2.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = _json.loads(resp.read())
        if data.get("result") == "success":
            rates = data.get("rates", {})
            usd_aud = rates.get("AUD")
            usd_bdt = rates.get("BDT")
            if usd_aud and usd_bdt:
                cache = {
                    "USD_AUD": round(usd_aud, 6),
                    "BDT_AUD": round(usd_aud / usd_bdt, 6) if usd_bdt else 1 / 86.15,
                    "ts": int(_time.time()),
                    "src": "open.er-api.com"
                }
                with _FX_LOCK:
                    _save_fx_cache(cache)
                return cache
    except Exception:
        pass
    return None

def get_fx_rates():
    """Return cached FX rates, refreshing in background if stale (>1 hour old)."""
    with _FX_LOCK:
        cache = _load_fx_cache()
    age = int(_time.time()) - cache.get("ts", 0)
    if age > 3600:
        _threading.Thread(target=_fetch_live_fx, daemon=True).start()
    return cache

# Initial fetch on startup
_fetch_live_fx()

FX_USD_TO_AUD = get_fx_rates().get("USD_AUD", 1.4275)
FX_BDT_TO_AUD = get_fx_rates().get("BDT_AUD", 1 / 86.15)

def fx_to_aud(amount, currency):
    """Convert a foreign-currency amount into AUD using live FX rates.
    Unknown currencies are treated as already AUD (no conversion) rather than guessed."""
    amount = amount or 0
    currency = (currency or "AUD").upper()
    if currency == "AUD":
        return amount
    rates = get_fx_rates()
    if currency == "USD":
        return amount * rates.get("USD_AUD", FX_USD_TO_AUD)
    if currency == "BDT":
        return amount * rates.get("BDT_AUD", FX_BDT_TO_AUD)
    return amount

app = Flask(__name__)
_sk = DATA / "secret.key"
if not _sk.exists():
    import secrets as _s; _sk.write_text(_s.token_hex(32))
app.secret_key = _sk.read_text().strip()

# ─── DB INIT ────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as c:
        c.executescript("""
        CREATE TABLE IF NOT EXISTS transactions (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            date      TEXT NOT NULL,
            type      TEXT NOT NULL CHECK(type IN ('income','expense')),
            category  TEXT NOT NULL,
            description TEXT,
            amount    REAL NOT NULL,
            currency  TEXT DEFAULT 'AUD',
            account   TEXT,
            notes     TEXT,
            created   TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS investments (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            name         TEXT NOT NULL,
            ticker       TEXT,
            inv_type     TEXT NOT NULL,
            units        REAL,
            cost_basis   REAL,
            current_value REAL,
            currency     TEXT DEFAULT 'AUD',
            country      TEXT DEFAULT 'Australia',
            notes        TEXT,
            user_id      INTEGER DEFAULT 1,
            created      TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS accounts (
            id       INTEGER PRIMARY KEY AUTOINCREMENT,
            name     TEXT NOT NULL,
            acc_type TEXT NOT NULL,
            balance  REAL DEFAULT 0,
            currency TEXT DEFAULT 'AUD',
            country  TEXT DEFAULT 'Australia',
            notes    TEXT
        );
        CREATE TABLE IF NOT EXISTS categories (
            id      INTEGER PRIMARY KEY AUTOINCREMENT,
            name    TEXT NOT NULL,
            type    TEXT NOT NULL DEFAULT 'expense',
            color   TEXT DEFAULT '#7c75f5',
            icon    TEXT DEFAULT '',
            created TEXT DEFAULT (datetime('now')),
            UNIQUE(name, type)
        );
        CREATE TABLE IF NOT EXISTS users (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            name          TEXT NOT NULL,
            emoji         TEXT DEFAULT '👤',
            color         TEXT DEFAULT '#1565c0',
            tfn           TEXT DEFAULT '',
            username      TEXT,
            password_hash TEXT,
            is_admin      INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS liabilities (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            name          TEXT NOT NULL,
            liab_type     TEXT NOT NULL DEFAULT 'Other',
            amount        REAL NOT NULL DEFAULT 0,
            currency      TEXT DEFAULT 'AUD',
            country       TEXT DEFAULT 'Australia',
            interest_rate REAL DEFAULT 0,
            notes         TEXT,
            user_id       INTEGER DEFAULT 1,
            created       TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS investment_snapshots (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            date        TEXT NOT NULL UNIQUE,
            total_value REAL NOT NULL,
            total_cost  REAL NOT NULL,
            created     TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS audit_log (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            ts        TEXT DEFAULT (datetime('now')),
            entity    TEXT NOT NULL,
            action    TEXT NOT NULL,
            entity_id INTEGER,
            summary   TEXT,
            details   TEXT
        );
        CREATE TABLE IF NOT EXISTS business_profile (
            id            INTEGER PRIMARY KEY CHECK (id = 1),
            business_name TEXT DEFAULT '',
            abn           TEXT DEFAULT '',
            tfn           TEXT DEFAULT '',
            gst_registered INTEGER DEFAULT 0,
            updated       TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS businesses (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id        INTEGER NOT NULL DEFAULT 1,
            business_name  TEXT DEFAULT '',
            abn            TEXT DEFAULT '',
            gst_registered INTEGER DEFAULT 0,
            created        TEXT DEFAULT (datetime('now')),
            updated        TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS recurring_templates (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            type        TEXT NOT NULL CHECK(type IN ('income','expense')),
            category    TEXT NOT NULL,
            description TEXT DEFAULT '',
            amount      REAL NOT NULL,
            currency    TEXT DEFAULT 'AUD',
            account     TEXT DEFAULT '',
            notes       TEXT DEFAULT '',
            user_id     INTEGER DEFAULT 1,
            frequency   TEXT NOT NULL CHECK(frequency IN ('weekly','monthly','yearly')),
            start_date  TEXT NOT NULL,
            end_date    TEXT,
            next_date   TEXT NOT NULL,
            active      INTEGER DEFAULT 1,
            created     TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS notes (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            title      TEXT DEFAULT '',
            content    TEXT DEFAULT '',
            color      TEXT DEFAULT '#a78bfa',
            pinned     INTEGER DEFAULT 0,
            created    TEXT DEFAULT (datetime('now')),
            updated    TEXT DEFAULT (datetime('now'))
        );
        """)

init_db()

def migrate_db():
    pending_audit = []
    with get_db() as c:
        for sql in ["ALTER TABLE transactions ADD COLUMN user_id INTEGER",
                    "ALTER TABLE accounts ADD COLUMN user_id INTEGER",
                    "ALTER TABLE investments ADD COLUMN user_id INTEGER",
                    "ALTER TABLE liabilities ADD COLUMN user_id INTEGER",
                    "ALTER TABLE users ADD COLUMN tfn TEXT DEFAULT ''",
                    "ALTER TABLE users ADD COLUMN username TEXT",
                    "ALTER TABLE users ADD COLUMN password_hash TEXT",
                    "ALTER TABLE users ADD COLUMN is_admin INTEGER DEFAULT 0"]:
            try: c.execute(sql)
            except: pass
        c.execute("UPDATE transactions SET user_id=1 WHERE user_id IS NULL")
        c.execute("UPDATE accounts SET user_id=1 WHERE user_id IS NULL")
        # Multi-user separation: every existing investment/liability predates per-user
        # ownership, so (per explicit user instruction) assign them all to user 1 (Pavel).
        # Other users can be reassigned individually later via the Owner field.
        c.execute("UPDATE investments SET user_id=1 WHERE user_id IS NULL")
        c.execute("UPDATE liabilities SET user_id=1 WHERE user_id IS NULL")
        # Create businesses table if not exists (for existing DBs) — supports multiple
        # businesses per user, replacing the old single-row business_profile.
        c.execute("""CREATE TABLE IF NOT EXISTS businesses (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id        INTEGER NOT NULL DEFAULT 1,
            business_name  TEXT DEFAULT '',
            abn            TEXT DEFAULT '',
            gst_registered INTEGER DEFAULT 0,
            created        TEXT DEFAULT (datetime('now')),
            updated        TEXT DEFAULT (datetime('now'))
        )""")
        # One-time safety-net migration: if the old single-row business_profile has real
        # data and hasn't been migrated yet (no businesses rows exist), carry it forward —
        # the business name/ABN into the new businesses table (owned by user 1), and the
        # TFN onto user 1's own record. Never fabricated — only copies what the user already entered.
        old_profile = c.execute("SELECT * FROM business_profile WHERE id=1").fetchone()
        if old_profile and not c.execute("SELECT id FROM businesses LIMIT 1").fetchone():
            if (old_profile["business_name"] or "").strip() or (old_profile["abn"] or "").strip():
                cur = c.execute(
                    "INSERT INTO businesses (user_id,business_name,abn,gst_registered) VALUES (1,?,?,?)",
                    (old_profile["business_name"] or "", old_profile["abn"] or "", old_profile["gst_registered"] or 0)
                )
                pending_audit.append(("business", "migration-add", cur.lastrowid, "Migrated legacy business profile to per-user businesses table"))
            if (old_profile["tfn"] or "").strip():
                c.execute("UPDATE users SET tfn=? WHERE id=1 AND (tfn IS NULL OR tfn='')", (old_profile["tfn"],))
        # Create notes table if not exists (for existing DBs)
        c.execute("""CREATE TABLE IF NOT EXISTS notes (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            title      TEXT DEFAULT '',
            content    TEXT DEFAULT '',
            color      TEXT DEFAULT '#a78bfa',
            pinned     INTEGER DEFAULT 0,
            created    TEXT DEFAULT (datetime('now')),
            updated    TEXT DEFAULT (datetime('now'))
        )""")
        # Link transactions to their recurring template (for tracking deletions)
        try:
            c.execute("ALTER TABLE transactions ADD COLUMN recurring_id INTEGER DEFAULT NULL")
        except Exception:
            pass
        # Create recurring_templates if not exists (for existing DBs)
        c.execute("""CREATE TABLE IF NOT EXISTS recurring_templates (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            type        TEXT NOT NULL CHECK(type IN ('income','expense')),
            category    TEXT NOT NULL,
            description TEXT DEFAULT '',
            amount      REAL NOT NULL,
            currency    TEXT DEFAULT 'AUD',
            account     TEXT DEFAULT '',
            notes       TEXT DEFAULT '',
            user_id     INTEGER DEFAULT 1,
            frequency   TEXT NOT NULL CHECK(frequency IN ('weekly','monthly','yearly')),
            start_date  TEXT NOT NULL,
            end_date    TEXT,
            next_date   TEXT NOT NULL,
            active      INTEGER DEFAULT 1,
            created     TEXT DEFAULT (datetime('now'))
        )""")
    if pending_audit:
        # log_audit() isn't defined yet at this point in module load order, so write
        # directly here (mirrors what log_audit() itself does) rather than calling it.
        with get_db() as c2:
            for entity, action, entity_id, summary in pending_audit:
                try:
                    c2.execute("INSERT INTO audit_log (entity,action,entity_id,summary) VALUES (?,?,?,?)",
                               (entity, action, entity_id, summary))
                except Exception:
                    pass
migrate_db()

def migrate_fix_known_data():
    """One-time, idempotent fixes for known-bad data already sitting in a live database
    (i.e. inserted before this fix existed). Each fix only runs if the bug signature is
    still present, so this is safe to leave in and re-run on every startup.

    Audit entries are collected and written via log_audit() AFTER the connection below
    is closed — log_audit() opens its own connection, and calling it while this one is
    still open causes a silent "database is locked" failure (SQLite only allows one
    writer connection at a time, and the failure was swallowed by log_audit()'s own
    try/except, so the audit trail for these migration actions went missing silently)."""
    pending_audit = []
    with get_db() as c:
        # Bug: SCHD/QQQ/VOO were seeded with a weighting bug (4968*0.55/0.20/0.05) and
        # had their USD values treated as AUD. Replace with real brokerage figures
        # (units, avg cost, current value) sourced from Pavel's actual portfolio.
        fixes = [
            ("SCHD", 2732.4,  133.67477, 3479.55, 3799.04),
            ("QQQ",  993.6,   1.83406,   920.71,  1137.98),
            ("VOO",  248.4,   0.04884,   32.83,   30.97),
        ]
        for name, bug_signature, units, cost_basis, current_value in fixes:
            row = c.execute("SELECT id, cost_basis FROM investments WHERE name=?", (name,)).fetchone()
            if row and abs(row["cost_basis"] - bug_signature) < 0.01:
                c.execute("UPDATE investments SET units=?, cost_basis=?, current_value=?, currency='USD' WHERE id=?",
                          (units, cost_basis, current_value, row["id"]))
                pending_audit.append(("investment", "migration-fix", row["id"], f"{name}: corrected weighting bug + FX currency tag"))

        # Add Superannuation as a tracked investment if it's not already there.
        if not c.execute("SELECT id FROM investments WHERE name='Superannuation'").fetchone():
            cur = c.execute(
                "INSERT INTO investments (name,ticker,inv_type,units,cost_basis,current_value,currency,country,notes) VALUES (?,?,?,?,?,?,?,?,?)",
                ("Superannuation", "", "Superannuation", 1, 22242, 22242, "AUD", "Australia", "Employer + personal contributions")
            )
            pending_audit.append(("investment", "migration-add", cur.lastrowid, "Superannuation balance added"))

        # Add Business Income category if missing (sole-trader/ABN income, separate from Salary/Freelance).
        if not c.execute("SELECT id FROM categories WHERE name='Business Income' AND type='income'").fetchone():
            c.execute("INSERT INTO categories (name,type,color,icon) VALUES (?,?,?,?)",
                      ("Business Income", "income", "#00c9a0", "🧾"))
            pending_audit.append(("category", "migration-add", None, "Business Income category added"))

    for entity, action, entity_id, summary in pending_audit:
        log_audit(entity, action, entity_id, summary)

# ─── AI ENGINE ──────────────────────────────────────────────────────────────
# RAG-style engine following NotebookLM architecture:
#   Provider fallback chain → grounded system prompts → financial data context

DEFAULT_PROVIDERS = [
    {"id":"p1","name":"OpenRouter (any model)","description":"Add your OpenRouter key in Settings","url":"https://openrouter.ai/api/v1","model":"openai/gpt-4o-mini","api_key":"","enabled":False},
    {"id":"p2","name":"OpenAI","description":"Add your OpenAI key in Settings","url":"https://api.openai.com/v1","model":"gpt-4o-mini","api_key":"","enabled":False},
    {"id":"p3","name":"Anthropic (Claude)","description":"Add your Anthropic key in Settings","url":"https://api.anthropic.com/v1","model":"claude-haiku-4-5-20251001","api_key":"","enabled":False},
    {"id":"p4","name":"Ollama (local)","description":"Free local LLM — install ollama.com","url":"http://localhost:11434","model":"qwen2.5:3b","api_key":"ollama","is_ollama":True,"enabled":True},
]

# Strict financial grounding prompt (modeled after NotebookLM's SYSTEM_PROMPT)
FINANCE_SYSTEM_PROMPT = """You are a strict personal finance assistant with access ONLY to the user's financial data provided below.

CRITICAL RULES — follow every one without exception:
- ONLY answer from the financial data provided in this message. Nothing else.
- You have NO internet access and NO market data — do not reference external prices, news, or live data.
- If the answer is NOT clearly derivable from the data: say "I don't have enough data to answer that. Add more transactions or accounts."
- Never fabricate numbers, dates, accounts, or transactions that aren't in the data.
- When giving recommendations, base them on the user's actual spending patterns and portfolio allocation.
- Use AUD currency for all amounts unless explicitly stated otherwise.
- Format responses in Markdown: **bold** for key numbers, bullet points for lists, > blockquotes for summaries.
- Be concise and direct — the user wants actionable insights, not lengthy explanations.
- When comparing investments, only reference the ones in the user's portfolio.
- Always show actual numbers from the data when available."""

def load_providers():
    if SETTINGS.exists():
        try:
            return json.loads(SETTINGS.read_text())["providers"]
        except:
            pass
    return DEFAULT_PROVIDERS

def save_providers(providers):
    SETTINGS.write_text(json.dumps({"providers": providers}, indent=2))

def _mask_key(key):
    if not key or len(key) <= 8:
        return "***"
    return key[:4] + "***" + key[-4:]

def get_provider_status():
    providers = load_providers()
    result = []
    for i, p in enumerate(providers):
        result.append({
            "id": p.get("id"),
            "name": p.get("name"),
            "enabled": p.get("enabled", True),
            "priority": i + 1,
            "is_active": False,
            "model": p.get("model", ""),
            "api_key_masked": _mask_key(p.get("api_key", "")),
        })
    # Mark the first enabled one as active
    for r in result:
        if r["enabled"]:
            r["is_active"] = True
            break
    return result

def detect_best_model():
    providers = [p for p in load_providers() if p.get("enabled")]
    if providers:
        p = providers[0]
        return f"{p.get('name', '?')}: {p.get('model', '?')}"
    # Fallback: check Ollama
    try:
        import urllib.request
        req = urllib.request.Request("http://localhost:11434/api/tags")
        with urllib.request.urlopen(req, timeout=3) as r:
            data = json.loads(r.read())
        models = [m["name"] for m in data.get("models", [])]
        # Preference order
        prefs = ["qwen2.5:3b", "llama3.2:latest", "gemma3:latest"]
        for pref in prefs:
            if pref in models:
                return f"Ollama: {pref}"
        if models:
            return f"Ollama: {models[0]}"
    except:
        pass
    return "No providers"

def _call_api(provider, messages):
    import urllib.request, json as _json
    url = provider["url"].rstrip("/")
    if not url.endswith("/chat/completions"):
        url += "/chat/completions"
    payload = {
        "model": provider["model"],
        "messages": messages,
        "max_tokens": 2048,
        "temperature": 0.3,
    }
    data = _json.dumps(payload).encode()
    req = urllib.request.Request(url, data=data, headers={
        "Content-Type": "application/json",
        "Authorization": f"Bearer {provider.get('api_key', '')}",
        "HTTP-Referer": "https://finance.yeahia.uk",
        "X-Title": "FinanceHub",
    })
    with urllib.request.urlopen(req, timeout=30) as r:
        resp = _json.loads(r.read())
    return resp["choices"][0]["message"]["content"]

def _call_ollama(provider, messages):
    import urllib.request, json as _json
    url = provider["url"].rstrip("/") + "/api/chat"
    payload = {"model": provider["model"], "messages": messages, "stream": False}
    data = _json.dumps(payload).encode()
    req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=45) as r:
        resp = _json.loads(r.read())
    return resp["message"]["content"]

def _llm(messages):
    providers = [p for p in load_providers() if p.get("enabled")]
    errors = []
    for p in providers:
        try:
            is_ollama = p.get("is_ollama") or "localhost:11434" in p.get("url", "")
            if is_ollama:
                return _call_ollama(p, messages)
            else:
                return _call_api(p, messages)
        except Exception as e:
            errors.append(f"{p.get('name', '?')}: {e}")
            continue
    # Hard fallback to local Ollama qwen2.5:3b
    try:
        import urllib.request, json as _json
        url = "http://localhost:11434/api/chat"
        payload = {"model": "qwen2.5:3b", "messages": messages, "stream": False}
        data = _json.dumps(payload).encode()
        req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=30) as r:
            resp = _json.loads(r.read())
        return resp["message"]["content"]
    except Exception as e:
        errors.append(f"Ollama fallback: {e}")
    return f"All AI providers failed. Errors:\n" + "\n".join(errors)

def build_finance_context():
    uid = _uid()
    summary = get_summary(uid)
    uf, up = _uf(uid)
    shf, shp = _uf_shared(uid)
    with get_db() as c:
        # Recent transactions (last 20)
        txns = [dict(r) for r in c.execute(
            f"SELECT date, type, category, description, amount, account FROM transactions{uf} ORDER BY date DESC LIMIT 20", up
        ).fetchall()]
        # All investments
        invs = [dict(r) for r in c.execute(
            f"SELECT name, ticker, inv_type, units, cost_basis, current_value, currency, country, notes FROM investments{shf} ORDER BY current_value DESC", shp
        ).fetchall()]
        # All accounts
        accs = [dict(r) for r in c.execute(
            f"SELECT name, acc_type, balance, currency, country FROM accounts{uf} ORDER BY balance DESC", up
        ).fetchall()]
        # Category breakdown
        cats = [dict(r) for r in c.execute(
            f"SELECT category, type, SUM(amount) as total, COUNT(*) as count FROM transactions{uf} GROUP BY category, type ORDER BY total DESC", up
        ).fetchall()]
        # Liabilities
        liabs = [dict(r) for r in c.execute(
            f"SELECT name, liab_type, amount, currency, country, interest_rate, notes FROM liabilities{shf} ORDER BY amount DESC", shp
        ).fetchall()]

    ctx = f"""=== FINANCIAL DATA SNAPSHOT ===

--- ACCOUNTS ({len(accs)} total) ---
Total Balance: A${summary['acc_total']:,.2f} (converted to AUD where needed; FX snapshot: 1 USD=A${FX_USD_TO_AUD}, 1 BDT=A${round(FX_BDT_TO_AUD,5)})
"""
    for a in accs:
        ctx += f"- {a['name']} ({a['acc_type']}, {a['country']}): {a['balance']:,.2f} {a['currency']}\n"

    ctx += f"\n--- LIABILITIES ({len(liabs)} total, A${summary['liab_total']:,.2f}) ---\n"
    for l in liabs:
        ctx += f"- {l['name']} ({l['liab_type']}, {l['country']}): {l['amount']:,.2f} {l['currency']}"
        if l.get('interest_rate'):
            ctx += f" @ {l['interest_rate']}%"
        ctx += "\n"
        if l.get('notes'):
            ctx += f"  Note: {l['notes']}\n"
    ctx += f"\nNet Worth (accounts + investments - liabilities): A${summary['net_worth']:,.2f}\n"

    ctx += f"\n--- INVESTMENTS ({len(invs)} positions, total A${summary['inv_total']:,.2f}) ---\n"
    ctx += f"Invested: A${summary.get('inv_cost', summary['inv_total'] - summary['inv_gain']):,.2f} | Gain/Loss: A${summary['inv_gain']:,.2f}\n"
    for inv in invs:
        gain = inv['current_value'] - inv['cost_basis']
        pct = (gain / inv['cost_basis'] * 100) if inv['cost_basis'] else 0
        ctx += f"- {inv['name']}"
        if inv['ticker']:
            ctx += f" ({inv['ticker']})"
        ctx += f": {inv['inv_type']}, {inv['units']} units, A${inv['current_value']:,.2f} (cost A${inv['cost_basis']:,.2f}, {'+'if gain>=0 else ''}{pct:.1f}%)\n"
        if inv.get('notes'):
            ctx += f"  Note: {inv['notes']}\n"

    ctx += f"\n--- MONTHLY SUMMARY (Current Month) ---\n"
    ctx += f"Income: A${summary['income_month']:,.2f}\n"
    ctx += f"Expenses: A${summary['expense_month']:,.2f}\n"
    ctx += f"Savings: A${summary['savings_month']:,.2f} (Rate: {summary['savings_rate']}%)\n"
    ctx += f"Net Worth: A${summary['net_worth']:,.2f}\n"

    ctx += f"\n--- CATEGORY BREAKDOWN (All Time) ---\n"
    for cat in cats:
        ctx += f"- {cat['category']} ({cat['type']}): A${cat['total']:,.2f} ({cat['count']} transactions)\n"

    ctx += f"\n--- RECENT TRANSACTIONS (Last {len(txns)}) ---\n"
    for t in txns:
        ctx += f"- {t['date']} | {t['type']} | {t['category']} | {t.get('description','')} | A${t['amount']:,.2f}\n"

    ctx += "\n=== END FINANCIAL DATA ==="
    return ctx

# ─── HELPERS ────────────────────────────────────────────────────────────────

def row_to_dict(row):
    return dict(row)

def log_audit(entity, action, entity_id=None, summary=""):
    """Record a change (add/edit/delete) to the audit log shown in Settings > Activity Log."""
    try:
        with get_db() as c:
            c.execute("INSERT INTO audit_log (entity,action,entity_id,summary) VALUES (?,?,?,?)",
                      (entity, action, entity_id, summary))
    except Exception:
        pass  # audit logging must never block the actual operation

def fy_bounds(d=None):
    """Australian financial year bounds (1 Jul - 30 Jun) for a given date."""
    d = d or datetime.date.today()
    if d.month >= 7:
        start = datetime.date(d.year, 7, 1)
        end   = datetime.date(d.year + 1, 6, 30)
        label = f"{d.year}-{str(d.year+1)[2:]}"
    else:
        start = datetime.date(d.year - 1, 7, 1)
        end   = datetime.date(d.year, 6, 30)
        label = f"{d.year-1}-{str(d.year)[2:]}"
    return start, end, label

def get_summary(uid=None):
    with get_db() as c:
        now = datetime.date.today()
        ym  = now.strftime("%Y-%m")
        uf  = " AND user_id=?" if uid else ""
        up  = (uid,) if uid else ()
        # Shared-aware filter for investments/liabilities: user_id=0 is the reserved
        # "Shared/Household" sentinel, visible under every individual user AND "All".
        # The Demo account is excluded from the Shared sentinel - it must only ever see
        # its own seeded rows, never the real household's shared investments/liabilities.
        shf = (" WHERE user_id=?" if uid else "") if uid == _DEMO_UID else (" WHERE (user_id=? OR user_id=0)" if uid else "")
        shp = (uid,) if uid else ()

        def ts(where, p=()):
            rows = c.execute(f"SELECT amount, currency FROM transactions WHERE {where}{uf}", tuple(p)+up).fetchall()
            return sum(fx_to_aud(r["amount"], r["currency"]) for r in rows)

        income_month  = ts("type='income' AND date LIKE ?",  (ym+"%",))
        expense_month = ts("type='expense' AND date LIKE ?", (ym+"%",))
        total_income  = ts("type='income'")
        total_expense = ts("type='expense'")
        savings       = income_month - expense_month
        savings_rate  = round(savings / income_month * 100, 1) if income_month > 0 else 0

        # Investments — convert each row to AUD using the manual FX snapshots before summing,
        # so USD/BDT positions aren't counted as if they were already AUD. Scoped to the
        # selected user (plus Shared) when a "View As" user is selected.
        inv_rows  = c.execute(f"SELECT current_value, cost_basis, currency FROM investments{shf}", shp).fetchall()
        inv_total = sum(fx_to_aud(r["current_value"], r["currency"]) for r in inv_rows)
        inv_cost  = sum(fx_to_aud(r["cost_basis"],    r["currency"]) for r in inv_rows)
        inv_gain  = inv_total - inv_cost

        # Accounts — same FX treatment (covers the BDT-denominated Bangladesh accounts).
        acc_sql = "SELECT balance, currency FROM accounts" + (" WHERE user_id=?" if uid else "")
        acc_rows = c.execute(acc_sql, (uid,) if uid else ()).fetchall()
        acc_total = sum(fx_to_aud(r["balance"], r["currency"]) for r in acc_rows)

        liab_rows  = c.execute(f"SELECT amount, currency FROM liabilities{shf}", shp).fetchall()
        liab_total = sum(fx_to_aud(r["amount"], r["currency"]) for r in liab_rows)

        net_worth = acc_total + inv_total - liab_total

        # Snapshot recording always reflects the TOTAL household investment value,
        # regardless of which user's view is being computed right now — otherwise
        # whichever user's page loads last each day would silently overwrite the shared
        # daily history used by the Dashboard's Investment Value Trend chart.
        # The public Demo account's seeded fake investments are excluded here so a demo
        # visit never pollutes Pavel's real net-worth history with random numbers.
        try:
            _all_inv = c.execute(
                "SELECT current_value, cost_basis, currency FROM investments WHERE user_id IS NULL OR user_id != ?",
                (_DEMO_UID or -1,)
            ).fetchall()
            _snap_total = sum(fx_to_aud(r["current_value"], r["currency"]) for r in _all_inv)
            _snap_cost  = sum(fx_to_aud(r["cost_basis"],    r["currency"]) for r in _all_inv)
            c.execute("INSERT OR REPLACE INTO investment_snapshots (date,total_value,total_cost) VALUES (?,?,?)",
                      (now.isoformat(), round(_snap_total,2), round(_snap_cost,2)))
        except Exception:
            pass

        monthly = []
        for i in range(5,-1,-1):
            mn  = (now.replace(day=1) - datetime.timedelta(days=30*i)).strftime("%Y-%m")
            monthly.append({"month": mn, "income": round(ts("type='income' AND date LIKE ?",(mn+"%",)),2),
                            "expense": round(ts("type='expense' AND date LIKE ?",(mn+"%",)),2)})

        # Weekly buckets — last 8 ISO weeks (Mon-Sun).
        weekly = []
        for i in range(7,-1,-1):
            wk_end   = now - datetime.timedelta(days=now.weekday()) + datetime.timedelta(days=6) - datetime.timedelta(weeks=i)
            wk_start = wk_end - datetime.timedelta(days=6)
            inc = ts("type='income' AND date>=? AND date<=?", (wk_start.isoformat(), wk_end.isoformat()))
            exp = ts("type='expense' AND date>=? AND date<=?", (wk_start.isoformat(), wk_end.isoformat()))
            weekly.append({"week": f"{wk_start.strftime('%d %b')}", "income": round(inc,2), "expense": round(exp,2)})

        # Yearly buckets — last 4 calendar years.
        yearly = []
        for y in range(now.year-3, now.year+1):
            inc = ts("type='income' AND date LIKE ?",  (f"{y}-%",))
            exp = ts("type='expense' AND date LIKE ?", (f"{y}-%",))
            yearly.append({"year": str(y), "income": round(inc,2), "expense": round(exp,2)})

        _cat_rows = c.execute(f"SELECT category, amount, currency FROM transactions WHERE type='expense' AND date LIKE ?{uf}",
                              (ym+"%",)+up).fetchall()
        _cat_totals = {}
        for r in _cat_rows:
            _cat_totals[r["category"]] = _cat_totals.get(r["category"], 0) + fx_to_aud(r["amount"], r["currency"])
        cats = sorted([{"category": k, "total": v} for k, v in _cat_totals.items()], key=lambda x: -x["total"])[:8]
        # Group by type using FX-converted AUD values — a raw SUM(current_value) here would
        # silently mix AUD/USD/BDT figures together (same bug class already fixed for inv_total/acc_total).
        inv_type_rows = c.execute(f"SELECT inv_type, current_value, currency FROM investments{shf}", shp).fetchall()
        _inv_type_totals = {}
        for r in inv_type_rows:
            _inv_type_totals[r["inv_type"]] = _inv_type_totals.get(r["inv_type"], 0) + fx_to_aud(r["current_value"], r["currency"])
        inv_types = sorted(_inv_type_totals.items(), key=lambda kv: -kv[1])
        recent = c.execute(f"SELECT * FROM transactions WHERE 1=1{uf} ORDER BY date DESC, id DESC LIMIT 10", up).fetchall()
        liabilities = c.execute(f"SELECT * FROM liabilities{shf} ORDER BY amount DESC", shp).fetchall()

        # Investment history buckets — derived only from real recorded snapshots (investment_snapshots).
        snaps = c.execute("SELECT date, total_value, total_cost FROM investment_snapshots ORDER BY date ASC").fetchall()
        snaps = [dict(r) for r in snaps]
        tracking_since = snaps[0]["date"] if snaps else now.isoformat()
        inv_weekly  = snaps[-8:]
        inv_monthly = {}
        for s in snaps:
            mkey = s["date"][:7]
            inv_monthly[mkey] = s  # latest snapshot in each month wins
        inv_monthly = list(inv_monthly.values())[-6:]
        inv_yearly = {}
        for s in snaps:
            ykey = s["date"][:4]
            inv_yearly[ykey] = s
        inv_yearly = list(inv_yearly.values())[-4:]

        return {
            "income_month":  round(income_month,2),  "expense_month": round(expense_month,2),
            "savings_month": round(savings,2),         "savings_rate":  savings_rate,
            "total_income":  round(total_income,2),    "total_expense": round(total_expense,2),
            "inv_total":     round(inv_total,2),        "inv_gain":      round(inv_gain,2),
            "inv_cost":      round(inv_cost,2),
            "acc_total":     round(acc_total,2),        "net_worth":     round(net_worth,2),
            "liab_total":    round(liab_total,2),       "liabilities":   [row_to_dict(r) for r in liabilities],
            "monthly":       monthly,         "weekly": weekly,        "yearly": yearly,
            "categories":    [{"name": r["category"], "value": round(r["total"],2)} for r in cats],
            "inv_types":     [{"name": name, "value": round(total,2)} for name, total in inv_types],
            "recent":        [row_to_dict(r) for r in recent],
            "inv_weekly":    inv_weekly, "inv_monthly": inv_monthly, "inv_yearly": inv_yearly,
            "tracking_since": tracking_since,
            "fx_usd_aud": FX_USD_TO_AUD, "fx_bdt_aud": FX_BDT_TO_AUD,
        }

# ─── AUTH ───────────────────────────────────────────────────────────────────
# Two distinct session keys:
#   session['user_id'] — the actual logged-in identity, set at /login.
#   session['view_uid'] — admin-only "View As" selection (None = All/household).
# _uid() keeps its old name/signature so every existing route/query below needs no change:
# admins get their View-As selection, everyone else is always forced to their own identity.

def _is_admin(): return bool(session.get('is_admin'))

def _uid():
    if _is_admin():
        return session.get('view_uid')
    return session.get('user_id')

def _can_set_user_id():
    """Whether a request may set an explicit user_id on a record it's creating/editing.
    True for unauthenticated API callers (e.g. the local pavel-ai-stack MCP integration —
    /api/* intentionally stays outside the login wall, so its behaviour must not change)
    and for admins. False for logged-in non-admin browser sessions — closes a gap where a
    tampered POST body could attribute a record to a different household member."""
    return (not session.get('user_id')) or _is_admin()

def _forbidden_if_not_owner(existing_uid, allow_shared=False):
    """For edit/delete-by-id endpoints: True if this request should be blocked because the
    record doesn't belong to the caller. Only enforced for logged-in non-admin sessions —
    unauthenticated API callers and admins are never blocked (same as today).
    The Shared (user_id=0) bypass below exists so real household members can edit each
    other's shared investments/liabilities — the public Demo account is deliberately never
    granted it, otherwise a visitor could PUT/DELETE a real shared record by guessing its id
    even though it never appears in Demo's own list view."""
    caller = session.get('user_id')
    if not caller or _is_admin():
        return False
    if allow_shared and existing_uid == 0 and caller != _DEMO_UID:
        return False
    return existing_uid is not None and existing_uid != caller

_AUTH_EXEMPT_ENDPOINTS = {'login_page', 'logout', 'demo_page', 'static'}

# ─── DEMO ACCOUNT ───────────────────────────────────────────────────────────
# A real but dedicated, public household member (username/password: demo/demo) so the
# /demo entry point can transparently log a visitor in and let them click through every
# real page (transactions, investments, accounts, liabilities, tax, AI) without ever
# hitting the login wall or touching real household data. _DEMO_UID is resolved once at
# startup (see _get_or_create_demo_user() near seed_users()) and is then treated as a
# special case everywhere Shared/household-wide data is aggregated, so a demo visit can
# never read or pollute Pavel's real figures.
DEMO_USERNAME = "demo"
DEMO_PASSWORD = "demo"
_DEMO_UID = None

@app.before_request
def _require_login():
    # API surface (used by the browser AND by the local pavel-ai-stack MCP integration,
    # which has no login mechanism) is intentionally left open — only HTML page routes
    # are gated. This is a deliberate, user-confirmed scope decision.
    if request.path.startswith('/api/') or request.path.startswith('/static/'):
        return
    if request.endpoint in _AUTH_EXEMPT_ENDPOINTS:
        return
    if not session.get('user_id'):
        return redirect(url_for('login_page', next=request.path))
    # The public Demo account must never actually keep anything a visitor adds, edits, or
    # deletes. Rather than special-case every CRUD endpoint, every full page load while
    # signed in as Demo wipes and reseeds fresh random data first - in-page edits still work
    # (the underlying API call succeeds normally), but the moment the page is refreshed,
    # revisited, or the visitor logs out and clicks Demo again, everything reverts to a
    # clean baseline. _DEMO_UID is None until startup seeding finishes, so this is also
    # naturally a no-op before that.
    if _DEMO_UID and session.get('user_id') == _DEMO_UID:
        _reset_demo_account_data(_DEMO_UID)

@app.context_processor
def inject_account():
    uid = session.get('user_id')
    if not uid:
        return {}
    with get_db() as c:
        row = c.execute("SELECT id, name, emoji, color, username, is_admin FROM users WHERE id=?", (uid,)).fetchone()
    return {"current_account": row_to_dict(row) if row else None}

# ─── LOGIN BRUTE-FORCE LOCKOUT ──────────────────────────────────────────────
# In-memory, per-username failed-attempt counter. Deliberately keyed by username
# (not IP) per spec: 5 wrong passwords in a row locks that account for 60s, even
# if the attacker switches IP/device. This also means the same wrong-credentials
# message is shown whether the username exists or not, so attempts can't be used
# to enumerate valid usernames. Single-process Flask app (no auto-reload, one
# worker) so a plain dict + lock is sufficient - no need for a DB table or Redis.
_LOGIN_LOCK = _threading.Lock()
_login_fails = {}  # username_lower -> {"count": int, "locked_until": monotonic-seconds}
_MAX_LOGIN_FAILS = 5
_LOGIN_LOCKOUT_SECONDS = 60

def _login_lock_status(username):
    """Remaining lockout seconds for this username (0 if not currently locked)."""
    if not username:
        return 0
    with _LOGIN_LOCK:
        rec = _login_fails.get(username)
        if not rec:
            return 0
        remaining = rec["locked_until"] - _time.monotonic()
        return int(remaining) + 1 if remaining > 0 else 0

def _login_record_fail(username):
    with _LOGIN_LOCK:
        rec = _login_fails.setdefault(username, {"count": 0, "locked_until": 0})
        rec["count"] += 1
        if rec["count"] >= _MAX_LOGIN_FAILS:
            rec["locked_until"] = _time.monotonic() + _LOGIN_LOCKOUT_SECONDS
            rec["count"] = 0

def _login_record_success(username):
    with _LOGIN_LOCK:
        _login_fails.pop(username, None)

@app.route("/login", methods=["GET", "POST"])
def login_page():
    if request.method == "GET" and session.get("user_id"):
        return redirect("/")
    error = None
    locked_seconds = 0
    if request.method == "POST":
        username = (request.form.get("username") or "").strip().lower()
        password = request.form.get("password") or ""
        locked_seconds = _login_lock_status(username)
        if locked_seconds > 0:
            error = f"Too many failed attempts. Try again in {locked_seconds}s."
        else:
            with get_db() as c:
                row = c.execute(
                    "SELECT id, password_hash, is_admin FROM users WHERE LOWER(username)=?", (username,)
                ).fetchone()
            if row and row["password_hash"] and check_password_hash(row["password_hash"], password):
                _login_record_success(username)
                session.clear()
                session["user_id"] = row["id"]
                session["is_admin"] = bool(row["is_admin"])
                nxt = request.form.get("next") or request.args.get("next") or "/"
                if not nxt.startswith("/"):
                    nxt = "/"
                return redirect(nxt)
            _login_record_fail(username)
            locked_seconds = _login_lock_status(username)
            error = (f"Too many failed attempts. Locked for {locked_seconds}s."
                     if locked_seconds > 0 else "Incorrect username or password.")
    return render_template("login.html", error=error, next=request.args.get("next", ""),
                            locked=locked_seconds > 0, lock_seconds=locked_seconds)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login_page"))

def _reset_demo_account_data(demo_uid):
    """Wipes and reseeds ONLY the Demo account's own rows with fresh random numbers. Called
    on every /demo visit AND on every subsequent page load of a Demo session (see
    _require_login() below) - so anything a visitor adds, edits, or deletes while playing
    with the dashboard is gone the moment they refresh, navigate, or log out and click Demo
    again. Always writes rows tagged with demo_uid - never user_id=0 (the real Shared
    sentinel) and never any other user's id - so this can never touch real household data."""
    rnd = random.Random()  # unseeded - different every visit
    today = datetime.date.today()
    with get_db() as c:
        c.execute("DELETE FROM transactions WHERE user_id=?", (demo_uid,))
        c.execute("DELETE FROM accounts WHERE user_id=?", (demo_uid,))
        c.execute("DELETE FROM investments WHERE user_id=?", (demo_uid,))
        c.execute("DELETE FROM liabilities WHERE user_id=?", (demo_uid,))
        c.execute("DELETE FROM recurring_templates WHERE user_id=?", (demo_uid,))
        c.execute("DELETE FROM businesses WHERE user_id=?", (demo_uid,))

        for name, typ, cur, country, lo, hi in [
            ("Everyday Checking",     "Checking",     "AUD", "Australia",   2000,  12000),
            ("High-Interest Savings", "Savings",      "AUD", "Australia",   8000,  28000),
            ("Term Deposit",          "Fixed Deposit","AUD", "Australia",  10000,  40000),
            ("Overseas Savings",      "Savings",      "BDT", "Bangladesh", 80000, 450000),
        ]:
            c.execute(
                "INSERT INTO accounts (name,acc_type,balance,currency,country,user_id) VALUES (?,?,?,?,?,?)",
                (name, typ, round(rnd.uniform(lo, hi), 2), cur, country, demo_uid)
            )

        for name, ticker, inv_type in [
            ("Demo Index Fund",   "VAS.AX", "AU ETF"),
            ("Demo Tech Stocks",  "TECH",   "Stocks"),
            ("Demo Crypto",       "BTC",    "Crypto"),
            ("Demo Term Deposit", None,     "Fixed Deposit"),
        ]:
            cost  = round(rnd.uniform(3000, 20000), 2)
            value = round(cost * rnd.uniform(0.85, 1.45), 2)
            c.execute(
                "INSERT INTO investments (name,ticker,inv_type,units,cost_basis,current_value,currency,country,user_id) "
                "VALUES (?,?,?,?,?,?,?,?,?)",
                (name, ticker, inv_type, round(rnd.uniform(10, 500), 2), cost, value, "AUD", "Australia", demo_uid)
            )

        c.execute(
            "INSERT INTO liabilities (name,liab_type,amount,currency,country,interest_rate,user_id) VALUES (?,?,?,?,?,?,?)",
            ("Demo Car Loan", "Car Loan", round(rnd.uniform(5000, 18000), 2), "AUD", "Australia",
             round(rnd.uniform(4, 9), 2), demo_uid)
        )

        expense_cats = ["Groceries","Rent","Utilities","Internet","Transport","Dining","Entertainment","Health","Insurance"]
        for i in range(45):
            d = today - datetime.timedelta(days=i)
            if d.day in (1, 15):
                c.execute(
                    "INSERT INTO transactions (date,type,category,description,amount,currency,account,user_id) "
                    "VALUES (?,?,?,?,?,?,?,?)",
                    (d.isoformat(), "income", "Salary", "Monthly Salary",
                     round(rnd.uniform(3200, 5200), 2), "AUD", "Everyday Checking", demo_uid)
                )
            if rnd.random() < 0.55:
                cat = rnd.choice(expense_cats)
                c.execute(
                    "INSERT INTO transactions (date,type,category,description,amount,currency,account,user_id) "
                    "VALUES (?,?,?,?,?,?,?,?)",
                    (d.isoformat(), "expense", cat, f"Demo {cat} expense",
                     round(rnd.uniform(15, 320), 2), "AUD", "Everyday Checking", demo_uid)
                )

@app.route("/demo")
def demo_page():
    """Public, no-login entry point: transparently logs the visitor into the dedicated,
    public Demo account and reseeds it with fresh random numbers, then redirects into the
    real dashboard. Every real page (transactions, investments, accounts, liabilities, tax,
    AI) works normally from here on, scoped to the Demo account's own seeded rows only -
    no real household data is ever read, written, or exposed."""
    _reset_demo_account_data(_DEMO_UID)
    session.clear()
    session['user_id'] = _DEMO_UID
    session['is_admin'] = False
    return redirect('/')

@app.route("/api/account", methods=["POST"])
def api_self_account_update():
    """Self-service username/password change. Despite the /api/ prefix this route is
    NOT exempt from login in practice - it explicitly requires an active session below,
    because it is a sensitive, identity-mutating action that the local MCP integration
    has no legitimate reason to call (unlike the read/write data endpoints, which stay
    open on purpose). Always requires the caller's current password, even for admins."""
    caller = session.get('user_id')
    if not caller:
        return jsonify({"ok": False, "error": "Login required"}), 401
    if caller == _DEMO_UID:
        return jsonify({"ok": False, "error": "Account changes are disabled in demo mode"}), 403
    d = request.json or {}
    current_password = d.get("current_password") or ""
    new_username = (d.get("username") or "").strip().lower()
    new_password = d.get("new_password") or ""
    with get_db() as c:
        row = c.execute("SELECT id, username, password_hash FROM users WHERE id=?", (caller,)).fetchone()
        if not row or not row["password_hash"] or not check_password_hash(row["password_hash"], current_password):
            return jsonify({"ok": False, "error": "Current password is incorrect"}), 403
        updates, params = [], []
        if new_username and new_username != (row["username"] or ""):
            if not re.match(r'^[a-z0-9._-]{3,32}$', new_username):
                return jsonify({"ok": False, "error": "Username must be 3-32 characters: letters, numbers, dots, underscores, hyphens"}), 400
            dup = c.execute("SELECT id FROM users WHERE LOWER(username)=? AND id!=?", (new_username, caller)).fetchone()
            if dup:
                return jsonify({"ok": False, "error": "Username already taken"}), 400
            updates.append("username=?"); params.append(new_username)
        if new_password:
            if len(new_password) < 6:
                return jsonify({"ok": False, "error": "New password must be at least 6 characters"}), 400
            updates.append("password_hash=?"); params.append(generate_password_hash(new_password))
        if updates:
            params.append(caller)
            c.execute(f"UPDATE users SET {','.join(updates)} WHERE id=?", params)
    return jsonify({"ok": True})

# ─── PAGES ──────────────────────────────────────────────────────────────────

def _uf(uid): return (" WHERE user_id=?", [uid]) if uid else ("", [])
# Shared-aware filter for investments/liabilities — user_id=0 is the reserved
# "Shared/Household" sentinel, visible under every individual user's view plus "All".
# Distinct from _uf() above, which stays strict (no Shared concept) for transactions/accounts.
def _uf_shared(uid):
    # The public Demo account must never see the real Shared/Household sentinel
    # (user_id=0) investments or liabilities - that is real household data. Demo falls
    # back to the strict, non-shared filter so it only ever sees its own seeded rows.
    if uid == _DEMO_UID:
        return _uf(uid)
    return (" WHERE (user_id=? OR user_id=0)", [uid]) if uid else ("", [])

@app.route("/")
def index():
    uid = _uid()
    summary = get_summary(uid)
    now = datetime.date.today().strftime("%B %Y")
    with get_db() as c:
        wh, p = _uf(uid)
        accounts = [row_to_dict(r) for r in c.execute(f"SELECT * FROM accounts{wh} ORDER BY country, name", p).fetchall()]
        users = [row_to_dict(r) for r in c.execute("SELECT id, name, emoji, color, tfn, username, is_admin FROM users ORDER BY id").fetchall()]
    return render_template("index.html", summary=summary, now=now, accounts=accounts, users=users, current_uid=uid)

@app.route("/Demo")
def demo_redirect():
    return redirect("/")

@app.route("/transactions")
def transactions_page():
    uid = _uid()
    # Auto-process any due recurring transactions. Never for Demo - process_recurring() has
    # no uid filter and would create REAL transactions for the real household's due recurring
    # templates; a demo visitor's own templates are wiped on every page load anyway, so this
    # call has no legitimate purpose for them.
    if uid != _DEMO_UID:
        process_recurring()
    with get_db() as c:
        wh, p = _uf(uid)
        rows = c.execute(f"SELECT * FROM transactions{wh} ORDER BY date DESC, id DESC", p).fetchall()
        cats = c.execute("SELECT * FROM categories ORDER BY type, name").fetchall()
        users = [row_to_dict(r) for r in c.execute("SELECT id, name, emoji, color, tfn, username, is_admin FROM users ORDER BY id").fetchall()]
        recurring = [row_to_dict(r) for r in c.execute(f"SELECT * FROM recurring_templates{wh} ORDER BY active DESC, next_date ASC", p).fetchall()]
    return render_template("transactions.html",
        transactions=[row_to_dict(r) for r in rows],
        categories=[row_to_dict(r) for r in cats],
        users=users, current_uid=uid,
        recurring=recurring,
        fx_usd_aud=FX_USD_TO_AUD, fx_bdt_aud=FX_BDT_TO_AUD,
        current_month=datetime.date.today().strftime("%Y-%m"))

@app.route("/investments")
def investments_page():
    uid = _uid()
    now_ym = datetime.date.today().strftime("%Y-%m")
    today = datetime.date.today().isoformat()
    with get_db() as c:
        wh, p = _uf_shared(uid)
        rows = c.execute(f"SELECT * FROM investments{wh} ORDER BY country, inv_type, name", p).fetchall()
        users = [row_to_dict(r) for r in c.execute("SELECT id, name, emoji, color, tfn, username, is_admin FROM users ORDER BY id").fetchall()]
        uf2 = " AND user_id=?" if uid else ""
        up2 = (uid,) if uid else ()
        # Property cash flow transactions (all) — scoped to the selected user like the rest of the page
        prop_txns = [row_to_dict(r) for r in c.execute(
            f"SELECT * FROM transactions WHERE category IN ('Rental Income','Property Maintenance','Property Repair'){uf2} ORDER BY date DESC", up2
        ).fetchall()]
        # Future property costs (date >= today) — expenses only, not income
        prop_future = [row_to_dict(r) for r in c.execute(
            f"SELECT * FROM transactions WHERE category IN ('Property Maintenance','Property Repair') AND date >= ?{uf2} ORDER BY date ASC",
            (today,)+up2
        ).fetchall()]
    # Compute property cash flow stats — amounts are in BDT (property is in Bangladesh)
    _bdt_to_aud = FX_BDT_TO_AUD
    prop_stats = {
        # BDT amounts (raw)
        'rent_total_bdt': sum(t['amount'] for t in prop_txns if t['category'] == 'Rental Income'),
        'maint_total_bdt': sum(t['amount'] for t in prop_txns if t['category'] == 'Property Maintenance'),
        'repair_total_bdt': sum(t['amount'] for t in prop_txns if t['category'] == 'Property Repair'),
        'rent_month_bdt': sum(t['amount'] for t in prop_txns if t['category'] == 'Rental Income' and t['date'][:7] == now_ym),
        'maint_month_bdt': sum(t['amount'] for t in prop_txns if t['category'] == 'Property Maintenance' and t['date'][:7] == now_ym),
        'future_total_bdt': sum(t['amount'] for t in prop_future),
        # AUD converted
        'rent_total': round(sum(t['amount'] for t in prop_txns if t['category'] == 'Rental Income') * _bdt_to_aud, 2),
        'maint_total': round(sum(t['amount'] for t in prop_txns if t['category'] == 'Property Maintenance') * _bdt_to_aud, 2),
        'repair_total': round(sum(t['amount'] for t in prop_txns if t['category'] == 'Property Repair') * _bdt_to_aud, 2),
        'rent_month': round(sum(t['amount'] for t in prop_txns if t['category'] == 'Rental Income' and t['date'][:7] == now_ym) * _bdt_to_aud, 2),
        'maint_month': round(sum(t['amount'] for t in prop_txns if t['category'] == 'Property Maintenance' and t['date'][:7] == now_ym) * _bdt_to_aud, 2),
        'future_total': round(sum(t['amount'] for t in prop_future) * _bdt_to_aud, 2),
        'future_count': len(prop_future),
    }
    return render_template("investments.html", investments=[row_to_dict(r) for r in rows],
                            fx_usd_aud=FX_USD_TO_AUD, fx_bdt_aud=FX_BDT_TO_AUD,
                            prop_txns=prop_txns, prop_future=prop_future,
                            prop_stats=prop_stats, now=now_ym, users=users, current_uid=uid)

@app.route("/accounts")
def accounts_page():
    uid = _uid()
    with get_db() as c:
        wh, p = _uf(uid)
        rows = c.execute(f"SELECT * FROM accounts{wh} ORDER BY country, name", p).fetchall()
        users = [row_to_dict(r) for r in c.execute("SELECT id, name, emoji, color, tfn, username, is_admin FROM users ORDER BY id").fetchall()]
    return render_template("accounts.html", accounts=[row_to_dict(r) for r in rows], users=users, current_uid=uid,
                            fx_usd_aud=FX_USD_TO_AUD, fx_bdt_aud=FX_BDT_TO_AUD)

@app.route("/liabilities")
def liabilities_page():
    uid = _uid()
    with get_db() as c:
        wh, p = _uf_shared(uid)
        rows = c.execute(f"SELECT * FROM liabilities{wh} ORDER BY amount DESC", p).fetchall()
        users = [row_to_dict(r) for r in c.execute("SELECT id, name, emoji, color, tfn, username, is_admin FROM users ORDER BY id").fetchall()]
    return render_template("liabilities.html", liabilities=[row_to_dict(r) for r in rows],
                            fx_usd_aud=FX_USD_TO_AUD, fx_bdt_aud=FX_BDT_TO_AUD,
                            users=users, current_uid=uid)

@app.route("/tax")
def tax_page():
    uid = _uid()
    uf = " AND user_id=?" if uid else ""
    up = (uid,) if uid else ()
    start, end, label = fy_bounds()
    with get_db() as c:
        income_rows = c.execute(
            f"SELECT category, SUM(amount) as total, COUNT(*) as n FROM transactions WHERE type='income' AND date>=? AND date<=?{uf} GROUP BY category ORDER BY total DESC",
            (start.isoformat(), end.isoformat())+up
        ).fetchall()
        expense_rows = c.execute(
            f"SELECT category, SUM(amount) as total, COUNT(*) as n FROM transactions WHERE type='expense' AND date>=? AND date<=?{uf} GROUP BY category ORDER BY total DESC",
            (start.isoformat(), end.isoformat())+up
        ).fetchall()
        business_income = c.execute(
            f"SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='income' AND category='Business Income' AND date>=? AND date<=?{uf}",
            (start.isoformat(), end.isoformat())+up
        ).fetchone()[0]
        shf, shp = _uf_shared(uid)
        investments = c.execute(f"SELECT * FROM investments{shf} ORDER BY country, inv_type, name", shp).fetchall()
        liabilities = c.execute(f"SELECT * FROM liabilities{shf} ORDER BY amount DESC", shp).fetchall()
        # Businesses (multiple per user) + TFN now live on the person, not a single shared
        # profile. "All" view shows every business across users; a selected user sees only theirs.
        biz_wh, biz_p = (" WHERE user_id=?", (uid,)) if uid else ("", ())
        businesses = [row_to_dict(r) for r in c.execute(f"SELECT * FROM businesses{biz_wh} ORDER BY id", biz_p).fetchall()]
        user_row = c.execute("SELECT id, name, emoji, color, tfn, username, is_admin FROM users WHERE id=?", (uid,)).fetchone() if uid else None
        users = [row_to_dict(r) for r in c.execute("SELECT id, name, emoji, color, tfn, username, is_admin FROM users ORDER BY id").fetchall()]
    tfn = (user_row["tfn"] or "") if user_row else ""
    tax_notes = _load_tax_notes()
    return render_template("tax.html",
        fy_label=label, fy_start=start, fy_end=end,
        income_rows=[row_to_dict(r) for r in income_rows],
        expense_rows=[row_to_dict(r) for r in expense_rows],
        business_income=round(business_income,2),
        investments=[row_to_dict(r) for r in investments],
        liabilities=[row_to_dict(r) for r in liabilities],
        businesses=businesses, tfn=tfn, tax_notes=tax_notes,
        users=users, current_uid=uid,
        fx_usd_aud=FX_USD_TO_AUD, fx_bdt_aud=FX_BDT_TO_AUD)

@app.route("/ai")
def ai_page():
    # AI chat is grounded in the caller's own data (safe) but also reads/writes a single
    # global chat_history.json shared by the whole household — a Demo visitor must never
    # see or pollute the real household's chat history, so Demo is kept off this page entirely.
    if _uid() == _DEMO_UID:
        return redirect('/')
    return render_template("ai.html")

@app.route("/settings")
def settings_page():
    # Settings shows the full real audit log + legacy business profile with no per-user
    # scoping at all - never let the public Demo account anywhere near it.
    if _uid() == _DEMO_UID:
        return redirect('/')
    providers = load_providers()
    with get_db() as c:
        audit_rows = c.execute("SELECT * FROM audit_log ORDER BY id DESC LIMIT 200").fetchall()
        profile_row = c.execute("SELECT * FROM business_profile WHERE id=1").fetchone()
    profile = row_to_dict(profile_row) if profile_row else {"business_name":"","abn":"","tfn":"","gst_registered":0}
    return render_template("settings.html", engines=providers,
                            audit_log=[row_to_dict(r) for r in audit_rows], profile=profile)

# ─── API: TRANSACTIONS ───────────────────────────────────────────────────────

@app.route("/api/transactions", methods=["GET"])
def api_get_transactions():
    q = request.args.get("q","").lower()
    type_f = request.args.get("type","")
    cat_f  = request.args.get("category","")
    month_f = request.args.get("month","")
    uid = _uid()
    with get_db() as c:
        sql = "SELECT * FROM transactions WHERE 1=1"
        params = []
        if uid:
            sql += " AND user_id=?"; params.append(uid)
        if q:
            sql += " AND (LOWER(description) LIKE ? OR LOWER(category) LIKE ? OR LOWER(notes) LIKE ?)"
            params += [f"%{q}%", f"%{q}%", f"%{q}%"]
        if type_f:
            sql += " AND type=?"; params.append(type_f)
        if cat_f:
            sql += " AND category=?"; params.append(cat_f)
        if month_f:
            sql += " AND date LIKE ?"; params.append(month_f+"%")
        sql += " ORDER BY date DESC, id DESC"
        rows = c.execute(sql, params).fetchall()
    return jsonify({"ok":True, "data":[row_to_dict(r) for r in rows]})

@app.route("/api/transactions", methods=["POST"])
def api_add_transaction():
    d = request.json
    required = ["date","type","category","amount"]
    if not all(d.get(k) for k in required):
        return jsonify({"ok":False,"error":"Missing required fields"})
    if d["type"] not in ("income","expense"):
        return jsonify({"ok":False,"error":"Type must be income or expense"})
    try:
        amount = float(d["amount"])
        if amount <= 0:
            return jsonify({"ok":False,"error":"Amount must be positive"})
    except:
        return jsonify({"ok":False,"error":"Invalid amount"})
    uid_in = d.get("user_id") if _can_set_user_id() else None
    uid = int(uid_in) if uid_in not in (None, "") else (_uid() or 1)
    with get_db() as c:
        cur = c.execute(
            "INSERT INTO transactions (date,type,category,description,amount,currency,account,notes,user_id) VALUES (?,?,?,?,?,?,?,?,?)",
            (d["date"], d["type"], d["category"], d.get("description",""), amount,
             d.get("currency","AUD"), d.get("account",""), d.get("notes",""), uid)
        )
        row = c.execute("SELECT * FROM transactions WHERE id=?", (cur.lastrowid,)).fetchone()
    log_audit("transaction", "add", cur.lastrowid, f"{d['type']} · {d['category']} · A${amount:,.2f}")
    return jsonify({"ok":True, "data": row_to_dict(row)})

@app.route("/api/transactions/<int:tid>", methods=["PUT"])
def api_update_transaction(tid):
    d = request.json
    with get_db() as c:
        existing = c.execute("SELECT user_id FROM transactions WHERE id=?", (tid,)).fetchone()
        if existing and _forbidden_if_not_owner(existing["user_id"]):
            return jsonify({"ok": False, "error": "Forbidden"}), 403
        c.execute(
            "UPDATE transactions SET date=?,type=?,category=?,description=?,amount=?,currency=?,account=?,notes=? WHERE id=?",
            (d["date"], d["type"], d["category"], d.get("description",""), float(d["amount"]),
             d.get("currency","AUD"), d.get("account",""), d.get("notes",""), tid)
        )
        row = c.execute("SELECT * FROM transactions WHERE id=?", (tid,)).fetchone()
    log_audit("transaction", "edit", tid, f"{d['type']} · {d['category']} · A${float(d['amount']):,.2f}")
    return jsonify({"ok":True, "data": row_to_dict(row) if row else None})

@app.route("/api/transactions/<int:tid>", methods=["DELETE"])
def api_delete_transaction(tid):
    with get_db() as c:
        row = c.execute("SELECT * FROM transactions WHERE id=?", (tid,)).fetchone()
        if row and _forbidden_if_not_owner(row["user_id"]):
            return jsonify({"ok": False, "error": "Forbidden"}), 403
        # If this transaction was created from a recurring template, deactivate it
        if row and row["recurring_id"]:
            c.execute("UPDATE recurring_templates SET active=0 WHERE id=?", (row["recurring_id"],))
        c.execute("DELETE FROM transactions WHERE id=?", (tid,))
    log_audit("transaction", "delete", tid, f"{row['category']} · A${row['amount']:,.2f}" if row else "")
    return jsonify({"ok":True})

# ─── API: INVESTMENTS ────────────────────────────────────────────────────────

@app.route("/api/investments", methods=["GET"])
def api_get_investments():
    # Optional ?user_id= filter (shared-aware); omitted = full household list, matching
    # the existing /api/transactions convention of explicit-param-only filtering.
    uid_f = request.args.get("user_id")
    caller = session.get('user_id')
    if caller and not _is_admin():
        uid_f = caller  # logged-in non-admin browser session: always own view, ignore any override
    with get_db() as c:
        if uid_f not in (None, ""):
            wh, p = _uf_shared(int(uid_f))
            rows = c.execute(f"SELECT * FROM investments{wh} ORDER BY country, inv_type, name", p).fetchall()
        else:
            rows = c.execute("SELECT * FROM investments ORDER BY country, inv_type, name").fetchall()
    return jsonify({"ok":True, "data":[row_to_dict(r) for r in rows]})

@app.route("/api/investments", methods=["POST"])
def api_add_investment():
    d = request.json
    uid_in = d.get("user_id") if _can_set_user_id() else None
    # 0 = Shared/Household — must be preserved, so check for None/"" explicitly rather than falsy.
    uid = int(uid_in) if uid_in not in (None, "") else (_uid() or 1)
    with get_db() as c:
        cur = c.execute(
            "INSERT INTO investments (name,ticker,inv_type,units,cost_basis,current_value,currency,country,notes,user_id) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (d["name"], d.get("ticker",""), d["inv_type"], float(d.get("units") or 0),
             float(d.get("cost_basis") or 0), float(d.get("current_value") or 0),
             d.get("currency","AUD"), d.get("country","Australia"), d.get("notes",""), uid)
        )
        row = c.execute("SELECT * FROM investments WHERE id=?", (cur.lastrowid,)).fetchone()
    log_audit("investment", "add", cur.lastrowid, f"{d['name']} · {d['inv_type']}")
    return jsonify({"ok":True, "data": row_to_dict(row)})

@app.route("/api/investments/<int:iid>", methods=["PUT"])
def api_update_investment(iid):
    d = request.json
    with get_db() as c:
        existing = c.execute("SELECT user_id FROM investments WHERE id=?", (iid,)).fetchone()
        if existing and _forbidden_if_not_owner(existing["user_id"], allow_shared=True):
            return jsonify({"ok": False, "error": "Forbidden"}), 403
        uid_in = d.get("user_id") if _can_set_user_id() else None
        uid = int(uid_in) if uid_in not in (None, "") else (existing["user_id"] if existing and existing["user_id"] is not None else 1)
        c.execute(
            "UPDATE investments SET name=?,ticker=?,inv_type=?,units=?,cost_basis=?,current_value=?,currency=?,country=?,notes=?,user_id=? WHERE id=?",
            (d["name"], d.get("ticker",""), d["inv_type"], float(d.get("units") or 0),
             float(d.get("cost_basis") or 0), float(d.get("current_value") or 0),
             d.get("currency","AUD"), d.get("country","Australia"), d.get("notes",""), uid, iid)
        )
        row = c.execute("SELECT * FROM investments WHERE id=?", (iid,)).fetchone()
    log_audit("investment", "edit", iid, f"{d['name']} · {d['inv_type']}")
    return jsonify({"ok":True, "data": row_to_dict(row) if row else None})

@app.route("/api/investments/<int:iid>", methods=["DELETE"])
def api_delete_investment(iid):
    with get_db() as c:
        row = c.execute("SELECT * FROM investments WHERE id=?", (iid,)).fetchone()
        if row and _forbidden_if_not_owner(row["user_id"], allow_shared=True):
            return jsonify({"ok": False, "error": "Forbidden"}), 403
        c.execute("DELETE FROM investments WHERE id=?", (iid,))
    log_audit("investment", "delete", iid, row["name"] if row else "")
    return jsonify({"ok":True})

# ─── API: ACCOUNTS ────────────────────────────────────────────────────────────

@app.route("/api/accounts", methods=["GET"])
def api_get_accounts():
    caller = session.get('user_id')
    with get_db() as c:
        if caller and not _is_admin():
            wh, p = _uf(caller)
            rows = c.execute(f"SELECT * FROM accounts{wh} ORDER BY country, name", p).fetchall()
        else:
            rows = c.execute("SELECT * FROM accounts ORDER BY country, name").fetchall()
    return jsonify({"ok":True, "data":[row_to_dict(r) for r in rows]})

@app.route("/api/accounts", methods=["POST"])
def api_add_account():
    d = request.json
    uid = _uid() or 1
    with get_db() as c:
        cur = c.execute(
            "INSERT INTO accounts (name,acc_type,balance,currency,country,notes,user_id) VALUES (?,?,?,?,?,?,?)",
            (d["name"], d["acc_type"], float(d.get("balance") or 0),
             d.get("currency","AUD"), d.get("country","Australia"), d.get("notes",""), uid)
        )
        row = c.execute("SELECT * FROM accounts WHERE id=?", (cur.lastrowid,)).fetchone()
    log_audit("account", "add", cur.lastrowid, f"{d['name']} · {d.get('currency','AUD')} {float(d.get('balance') or 0):,.2f}")
    return jsonify({"ok":True, "data": row_to_dict(row)})

@app.route("/api/accounts/<int:aid>", methods=["PUT"])
def api_update_account(aid):
    d = request.json
    with get_db() as c:
        old = c.execute("SELECT balance,currency,user_id FROM accounts WHERE id=?", (aid,)).fetchone()
        if old and _forbidden_if_not_owner(old["user_id"]):
            return jsonify({"ok": False, "error": "Forbidden"}), 403
        c.execute(
            "UPDATE accounts SET name=?,acc_type=?,balance=?,currency=?,country=?,notes=? WHERE id=?",
            (d["name"], d["acc_type"], float(d.get("balance") or 0),
             d.get("currency","AUD"), d.get("country","Australia"), d.get("notes",""), aid)
        )
        row = c.execute("SELECT * FROM accounts WHERE id=?", (aid,)).fetchone()
    old_bal = f"{old['currency']} {old['balance']:,.2f}" if old else "?"
    new_bal = f"{d.get('currency','AUD')} {float(d.get('balance') or 0):,.2f}"
    log_audit("account", "edit", aid, f"{d['name']} · {old_bal} → {new_bal}")
    return jsonify({"ok":True, "data": row_to_dict(row) if row else None})

@app.route("/api/accounts/<int:aid>", methods=["DELETE"])
def api_delete_account(aid):
    with get_db() as c:
        row = c.execute("SELECT * FROM accounts WHERE id=?", (aid,)).fetchone()
        if row and _forbidden_if_not_owner(row["user_id"]):
            return jsonify({"ok": False, "error": "Forbidden"}), 403
        c.execute("DELETE FROM accounts WHERE id=?", (aid,))
    log_audit("account", "delete", aid, row["name"] if row else "")
    return jsonify({"ok":True})

# ─── API: LIABILITIES ────────────────────────────────────────────────────────

@app.route("/api/liabilities", methods=["GET"])
def api_get_liabilities():
    # Optional ?user_id= filter (shared-aware); omitted = full household list.
    uid_f = request.args.get("user_id")
    caller = session.get('user_id')
    if caller and not _is_admin():
        uid_f = caller  # logged-in non-admin browser session: always own view, ignore any override
    with get_db() as c:
        if uid_f not in (None, ""):
            wh, p = _uf_shared(int(uid_f))
            rows = c.execute(f"SELECT * FROM liabilities{wh} ORDER BY amount DESC", p).fetchall()
        else:
            rows = c.execute("SELECT * FROM liabilities ORDER BY amount DESC").fetchall()
    return jsonify({"ok":True, "data":[row_to_dict(r) for r in rows]})

@app.route("/api/liabilities", methods=["POST"])
def api_add_liability():
    d = request.json
    if not d.get("name"):
        return jsonify({"ok":False,"error":"Name is required"})
    uid_in = d.get("user_id") if _can_set_user_id() else None
    uid = int(uid_in) if uid_in not in (None, "") else (_uid() or 1)
    with get_db() as c:
        cur = c.execute(
            "INSERT INTO liabilities (name,liab_type,amount,currency,country,interest_rate,notes,user_id) VALUES (?,?,?,?,?,?,?,?)",
            (d["name"], d.get("liab_type","Other"), float(d.get("amount") or 0),
             d.get("currency","AUD"), d.get("country","Australia"),
             float(d.get("interest_rate") or 0), d.get("notes",""), uid)
        )
        row = c.execute("SELECT * FROM liabilities WHERE id=?", (cur.lastrowid,)).fetchone()
    log_audit("liability", "add", cur.lastrowid, f"{d['name']} · {d.get('currency','AUD')} {float(d.get('amount') or 0):,.2f}")
    return jsonify({"ok":True, "data": row_to_dict(row)})

@app.route("/api/liabilities/<int:lid>", methods=["PUT"])
def api_update_liability(lid):
    d = request.json
    with get_db() as c:
        existing = c.execute("SELECT user_id FROM liabilities WHERE id=?", (lid,)).fetchone()
        if existing and _forbidden_if_not_owner(existing["user_id"], allow_shared=True):
            return jsonify({"ok": False, "error": "Forbidden"}), 403
        uid_in = d.get("user_id") if _can_set_user_id() else None
        uid = int(uid_in) if uid_in not in (None, "") else (existing["user_id"] if existing and existing["user_id"] is not None else 1)
        c.execute(
            "UPDATE liabilities SET name=?,liab_type=?,amount=?,currency=?,country=?,interest_rate=?,notes=?,user_id=? WHERE id=?",
            (d["name"], d.get("liab_type","Other"), float(d.get("amount") or 0),
             d.get("currency","AUD"), d.get("country","Australia"),
             float(d.get("interest_rate") or 0), d.get("notes",""), uid, lid)
        )
        row = c.execute("SELECT * FROM liabilities WHERE id=?", (lid,)).fetchone()
    log_audit("liability", "edit", lid, f"{d['name']} · {d.get('currency','AUD')} {float(d.get('amount') or 0):,.2f}")
    return jsonify({"ok":True, "data": row_to_dict(row) if row else None})

@app.route("/api/liabilities/<int:lid>", methods=["DELETE"])
def api_delete_liability(lid):
    with get_db() as c:
        row = c.execute("SELECT * FROM liabilities WHERE id=?", (lid,)).fetchone()
        if row and _forbidden_if_not_owner(row["user_id"], allow_shared=True):
            return jsonify({"ok": False, "error": "Forbidden"}), 403
        c.execute("DELETE FROM liabilities WHERE id=?", (lid,))
    log_audit("liability", "delete", lid, row["name"] if row else "")
    return jsonify({"ok":True})

# ─── API: AUDIT / ACTIVITY LOG ───────────────────────────────────────────────

@app.route("/api/audit-log", methods=["GET"])
def api_get_audit_log():
    # This is the real, un-scoped change history of every edit ever made - was previously
    # reachable with zero login. Pre-existing bug, unrelated to the Demo feature, fixed here.
    _err = _require_session_json()
    if _err: return _err
    limit = int(request.args.get("limit", 200))
    with get_db() as c:
        rows = c.execute("SELECT * FROM audit_log ORDER BY id DESC LIMIT ?", (limit,)).fetchall()
    return jsonify({"ok":True, "data":[row_to_dict(r) for r in rows]})

# ─── API: BUSINESS PROFILE (ABN / TFN) ───────────────────────────────────────
# Fields are entered directly by the user in Settings and never pre-filled or
# requested by the assistant — stored locally only, never sent to any AI provider.

@app.route("/api/profile", methods=["GET"])
def api_get_profile():
    # Real ABN/TFN, previously reachable with zero login - pre-existing bug, unrelated to
    # the Demo feature, fixed here per the standing rule that real ABN/TFN must never be
    # exposed to anyone but the logged-in household.
    _err = _require_session_json()
    if _err: return _err
    with get_db() as c:
        row = c.execute("SELECT * FROM business_profile WHERE id=1").fetchone()
    if not row:
        return jsonify({"ok":True, "data":{"business_name":"","abn":"","tfn":"","gst_registered":0}})
    return jsonify({"ok":True, "data": row_to_dict(row)})

@app.route("/api/profile", methods=["POST"])
def api_save_profile():
    _err = _require_session_json()
    if _err: return _err
    d = request.json
    with get_db() as c:
        c.execute("""
            INSERT INTO business_profile (id,business_name,abn,tfn,gst_registered,updated)
            VALUES (1,?,?,?,?,datetime('now'))
            ON CONFLICT(id) DO UPDATE SET
                business_name=excluded.business_name, abn=excluded.abn,
                tfn=excluded.tfn, gst_registered=excluded.gst_registered, updated=excluded.updated
        """, (d.get("business_name",""), d.get("abn",""), d.get("tfn",""), int(bool(d.get("gst_registered")))))
    log_audit("profile", "edit", 1, "Business profile updated")
    return jsonify({"ok":True})

# ─── API: BUSINESSES (multiple per user — replaces the single business_profile) ─────
# TFN lives on the person (users.tfn below), not on a business — a person can run
# several businesses (each with its own ABN/GST status) under one TFN.

@app.route("/api/businesses", methods=["GET"])
def api_get_businesses():
    # Returns real ABNs - previously reachable with zero login (and would return every
    # real business across the household when called with no params). Pre-existing bug,
    # unrelated to the Demo feature, fixed here.
    _err = _require_session_json()
    if _err: return _err
    uid_f = request.args.get("user_id")
    caller = session.get('user_id')
    if caller and not _is_admin():
        uid_f = caller  # logged-in non-admin browser session: always own view, ignore any override
    with get_db() as c:
        if uid_f not in (None, ""):
            rows = c.execute("SELECT * FROM businesses WHERE user_id=? ORDER BY id", (int(uid_f),)).fetchall()
        else:
            rows = c.execute("SELECT * FROM businesses ORDER BY user_id, id").fetchall()
    return jsonify({"ok":True, "data":[row_to_dict(r) for r in rows]})

@app.route("/api/businesses", methods=["POST"])
def api_add_business():
    d = request.json
    if not d.get("business_name"):
        return jsonify({"ok":False,"error":"Business name is required"})
    uid_in = d.get("user_id") if _can_set_user_id() else None
    uid = int(uid_in) if uid_in not in (None, "") else (_uid() or 1)
    with get_db() as c:
        cur = c.execute(
            "INSERT INTO businesses (user_id,business_name,abn,gst_registered,updated) VALUES (?,?,?,?,datetime('now'))",
            (uid, d["business_name"].strip(), d.get("abn","").strip(), int(bool(d.get("gst_registered"))))
        )
        row = c.execute("SELECT * FROM businesses WHERE id=?", (cur.lastrowid,)).fetchone()
    log_audit("business", "add", cur.lastrowid, d["business_name"])
    return jsonify({"ok":True, "data": row_to_dict(row)})

@app.route("/api/businesses/<int:bid>", methods=["PUT"])
def api_update_business(bid):
    d = request.json
    with get_db() as c:
        existing = c.execute("SELECT user_id FROM businesses WHERE id=?", (bid,)).fetchone()
        if existing and _forbidden_if_not_owner(existing["user_id"]):
            return jsonify({"ok": False, "error": "Forbidden"}), 403
        uid_in = d.get("user_id") if _can_set_user_id() else None
        uid = int(uid_in) if uid_in not in (None, "") else (existing["user_id"] if existing else (_uid() or 1))
        c.execute(
            "UPDATE businesses SET business_name=?,abn=?,gst_registered=?,user_id=?,updated=datetime('now') WHERE id=?",
            (d.get("business_name","").strip(), d.get("abn","").strip(), int(bool(d.get("gst_registered"))), uid, bid)
        )
        row = c.execute("SELECT * FROM businesses WHERE id=?", (bid,)).fetchone()
    log_audit("business", "edit", bid, d.get("business_name",""))
    return jsonify({"ok":True, "data": row_to_dict(row) if row else None})

@app.route("/api/businesses/<int:bid>", methods=["DELETE"])
def api_delete_business(bid):
    with get_db() as c:
        row = c.execute("SELECT * FROM businesses WHERE id=?", (bid,)).fetchone()
        if row and _forbidden_if_not_owner(row["user_id"]):
            return jsonify({"ok": False, "error": "Forbidden"}), 403
        c.execute("DELETE FROM businesses WHERE id=?", (bid,))
    log_audit("business", "delete", bid, row["business_name"] if row else "")
    return jsonify({"ok":True})

# ─── API: USER TFN (stored on the person, never pre-filled or requested by the assistant) ──

@app.route("/api/users/<int:uid>/tfn", methods=["PUT"])
def api_update_user_tfn(uid):
    caller = session.get('user_id')
    if caller and not _is_admin() and caller != uid:
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    d = request.json
    with get_db() as c:
        c.execute("UPDATE users SET tfn=? WHERE id=?", (d.get("tfn","").strip(), uid))
        row = c.execute("SELECT id, name, emoji, color, tfn, username, is_admin FROM users WHERE id=?", (uid,)).fetchone()
    log_audit("user", "edit-tfn", uid, "TFN updated")
    return jsonify({"ok":True, "data": row_to_dict(row) if row else None})

# ─── API: TAX NOTES (previous years — free text, user-supplied) ─────────────

def _load_tax_notes():
    if TAX_NOTES_FILE.exists():
        try:
            return json.loads(TAX_NOTES_FILE.read_text())
        except Exception:
            pass
    return {}

@app.route("/api/tax/notes", methods=["GET"])
def api_get_tax_notes():
    _err = _require_session_json()
    if _err: return _err
    return jsonify({"ok":True, "data": _load_tax_notes()})

@app.route("/api/tax/notes", methods=["POST"])
def api_save_tax_notes():
    _err = _require_session_json()
    if _err: return _err
    d = request.json
    fy = d.get("fy","")
    if not fy:
        return jsonify({"ok":False,"error":"Financial year label required"})
    notes = _load_tax_notes()
    notes[fy] = d.get("notes","")
    TAX_NOTES_FILE.write_text(json.dumps(notes, indent=2))
    return jsonify({"ok":True})

# ─── API: SUMMARY ────────────────────────────────────────────────────────────

@app.route("/api/summary")
def api_summary():
    return jsonify(get_summary(_uid()))

# ─── API: AI CHAT ────────────────────────────────────────────────────────────
# NotebookLM-style: grounded in financial data, provider fallback chain, persistent history

CHAT_HISTORY = DATA / "chat_history.json"

def _load_chat_history():
    if CHAT_HISTORY.exists():
        try:
            return json.loads(CHAT_HISTORY.read_text())
        except:
            pass
    return []

def _save_chat_history(history):
    CHAT_HISTORY.write_text(json.dumps(history, indent=2))

@app.route("/api/chat", methods=["POST"])
def api_chat():
    # Block the public Demo account specifically (not anonymous callers — the local
    # pavel-ai-stack MCP's fin_chat tool calls this with no session and must keep working).
    # Chat history below is a single global file shared by the whole household, so letting
    # Demo chat would mix fake demo turns into Pavel's real history and feed Pavel's real
    # past Q&A back to a demo visitor as conversation context.
    if _uid() == _DEMO_UID:
        return jsonify({"ok": False, "error": "AI chat is disabled in demo mode"}), 403
    d = request.json
    user_msg = d.get("message", "").strip()
    if not user_msg:
        return jsonify({"ok": False, "error": "No message"})

    # Load persistent history
    chat_history = _load_chat_history()

    # Build grounded financial context (RAG-style)
    finance_data = build_finance_context()

    # Build messages array: system prompt + data context + history + user question
    messages = [
        {"role": "system", "content": FINANCE_SYSTEM_PROMPT},
        {"role": "system", "content": f"USER'S FINANCIAL DATA:\n\n{finance_data}"},
    ]

    # Inject conversation history (last 10 turns as context)
    recent = chat_history[-10:] if len(chat_history) > 10 else chat_history
    for turn in recent:
        messages.append({"role": "user", "content": turn["q"]})
        messages.append({"role": "assistant", "content": turn["a"]})

    # Add current question
    messages.append({"role": "user", "content": user_msg})

    # Call LLM with provider fallback chain
    reply = _llm(messages)

    # Persist to history
    chat_history.append({
        "q": user_msg,
        "a": reply,
        "ts": datetime.datetime.now().isoformat()
    })
    # Keep last 100 turns
    if len(chat_history) > 100:
        chat_history = chat_history[-100:]
    _save_chat_history(chat_history)

    return jsonify({"ok": True, "reply": reply})

@app.route("/api/chat/history", methods=["GET"])
def api_chat_history():
    # Same global-file concern as /api/chat above — Demo must never read Pavel's real history.
    if _uid() == _DEMO_UID:
        return jsonify({"ok": True, "history": []})
    return jsonify({"ok": True, "history": _load_chat_history()})

@app.route("/api/chat/history", methods=["DELETE"])
def api_clear_chat_history():
    # Never let the public Demo account wipe the real household's chat history.
    if _uid() == _DEMO_UID:
        return jsonify({"ok": False, "error": "Not available in demo mode"}), 403
    _save_chat_history([])
    return jsonify({"ok": True})

# ─── API: PROVIDERS ──────────────────────────────────────────────────────────

@app.route("/api/providers", methods=["GET"])
def api_get_providers():
    providers = load_providers()
    # Mask API keys before sending to browser (NotebookLM pattern)
    masked = []
    for p in providers:
        pc = dict(p)
        pc["api_key"] = _mask_key(pc.get("api_key", ""))
        masked.append(pc)
    return jsonify({"ok": True, "providers": masked})

@app.route("/api/providers", methods=["POST"])
def api_save_providers():
    d = request.json
    new_providers = d.get("providers", [])
    # If a key contains ***, restore the original from disk (prevent accidental erasure)
    if SETTINGS.exists():
        try:
            old_providers = json.loads(SETTINGS.read_text())["providers"]
            old_keys = {p["id"]: p.get("api_key", "") for p in old_providers}
            for p in new_providers:
                if "***" in p.get("api_key", "") and p["id"] in old_keys:
                    p["api_key"] = old_keys[p["id"]]
        except:
            pass
    save_providers(new_providers)
    return jsonify({"ok": True})

@app.route("/api/status")
def api_status():
    providers = load_providers()
    enabled = [p for p in providers if p.get("enabled")]
    active_model = detect_best_model()
    is_ollama = not enabled  # If no API providers, it's Ollama fallback

    # Check each provider's health
    provider_statuses = get_provider_status()

    return jsonify({
        "engine": "free_api" if enabled else "ollama",
        "active": active_model,
        "ollama": is_ollama,
        "providers": provider_statuses,
        "total": len(providers),
        "enabled_count": len(enabled),
    })

# ─── EXPORT ──────────────────────────────────────────────────────────────────

@app.route("/api/export/<fmt>")
def api_export(fmt):
    export_type = request.args.get("type","transactions")
    categories = request.args.get("categories","").strip()
    date_from = request.args.get("date_from","").strip()
    date_to = request.args.get("date_to","").strip()
    tx_type = request.args.get("tx_type","").strip()

    uid = _uid()
    with get_db() as c:
        if export_type == "transactions":
            sql = "SELECT * FROM transactions WHERE 1=1"
            params = []
            if uid:
                sql += " AND user_id=?"; params.append(uid)
            if categories:
                cat_list = [x.strip() for x in categories.split(",") if x.strip()]
                placeholders = ",".join(["?" for _ in cat_list])
                sql += f" AND category IN ({placeholders})"
                params += cat_list
            if date_from:
                sql += " AND date >= ?"
                params.append(date_from)
            if date_to:
                sql += " AND date <= ?"
                params.append(date_to)
            if tx_type:
                sql += " AND type = ?"
                params.append(tx_type)
            sql += " ORDER BY date DESC"
            rows = [row_to_dict(r) for r in c.execute(sql, params).fetchall()]
            headers = ["id","date","type","category","description","amount","currency","account","notes","created"]
        elif export_type == "investments":
            wh, p = _uf_shared(uid)
            rows = [row_to_dict(r) for r in c.execute(f"SELECT * FROM investments{wh} ORDER BY country, inv_type, name", p).fetchall()]
            headers = ["id","name","ticker","inv_type","units","cost_basis","current_value","currency","country","notes"]
        elif export_type == "accounts":
            wh, p = _uf(uid)
            rows = [row_to_dict(r) for r in c.execute(f"SELECT * FROM accounts{wh} ORDER BY country, name", p).fetchall()]
            headers = ["id","name","acc_type","balance","currency","country","notes"]
        else:
            rows, headers = [], []

    if fmt == "csv":
        import csv
        buf = io.StringIO()
        w = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
        w.writeheader(); w.writerows(rows)
        return send_file(io.BytesIO(buf.getvalue().encode()), download_name=f"{export_type}.csv", as_attachment=True, mimetype="text/csv")

    elif fmt == "xlsx":
        try:
            import openpyxl
        except ImportError:
            return jsonify({"ok":False,"error":"openpyxl not installed. Run: pip install openpyxl"}), 400
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = export_type.capitalize()
        from openpyxl.styles import Font, PatternFill, Alignment
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="1a1a2e")
        ws.append(headers)
        for cell in ws[1]:
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append([row.get(h,"") for h in headers])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or ""))+4, 14)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, download_name=f"{export_type}.xlsx", as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    elif fmt == "pdf":
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
        except ImportError:
            return jsonify({"ok":False,"error":"reportlab not installed. Run: pip install reportlab"}), 400
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = [Paragraph(f"Personal Finance — {export_type.capitalize()}", styles["Title"]), Spacer(1,12)]
        table_data = [headers] + [[str(row.get(h,"")) for h in headers] for row in rows]
        t = Table(table_data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1a1a2e")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("GRID",(0,0),(-1,-1),0.5,colors.grey),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f5f5f5")]),
            ("FONTSIZE",(0,0),(-1,-1),8),
        ]))
        elements.append(t)
        doc.build(elements)
        buf.seek(0)
        return send_file(buf, download_name=f"{export_type}.pdf", as_attachment=True, mimetype="application/pdf")

    return jsonify({"ok":False,"error":"Unknown format"}), 400

# ─── SEED DATA (first run) ───────────────────────────────────────────────────

def seed_data():
    with get_db() as c:
        count = c.execute("SELECT COUNT(*) FROM accounts").fetchone()[0]
        if count > 0:
            return
        # Accounts from Pavel's profile
        accounts_seed = [
            ("CommBank NetSaver", "Savings", 18000, "AUD", "Australia", "Emergency Fund"),
            ("ING Fixed Deposit",  "Fixed Deposit", 10000, "AUD", "Australia", "6 months, locked"),
            ("ANZ Account",        "Checking", 200,   "AUD", "Australia", ""),
            ("Westpac Account",    "Checking", 140,   "AUD", "Australia", ""),
            ("BD Bank",            "Savings",  100000, "BDT", "Bangladesh", "≈ A$1,390 at seed-time FX (1 BDT≈A$0.0139)"),
        ]
        for a in accounts_seed:
            c.execute("INSERT OR IGNORE INTO accounts (name,acc_type,balance,currency,country,notes) VALUES (?,?,?,?,?,?)", a)

        # Investments — real figures sourced from Pavel's brokerage/ChatGPT financial profile (22 Jun 2026).
        # US ETF cost_basis = avg_cost * units; current_value as reported. Currency kept native (USD/AUD)
        # and converted to AUD only at summary/aggregate time via fx_to_aud().
        investments_seed = [
            ("SCHD",  "SCHD",  "US ETF",         133.67477, 3479.55,  3799.04,  "USD","USA",   "Income/Dividends"),
            ("QQQ",   "QQQ",   "US ETF",         1.83406,   920.71,   1137.98,  "USD","USA",   "Growth"),
            ("VOO",   "VOO",   "US ETF",         0.04884,   32.83,    30.97,    "USD","USA",   "Broad market"),
            ("ETHI",  "ETHI",  "Australian ETF", 30,        499.70,   499.70,   "AUD","Australia",""),
            ("IEM",   "IEM",   "Australian ETF", 5,         417.00,   417.00,   "AUD","Australia",""),
            ("NDQ",   "NDQ",   "Australian ETF", 3,         171.89,   171.89,   "AUD","Australia",""),
            ("Bangladesh Property","","Property",1,         87040,    87040,    "AUD","Bangladesh","58L BDT total; 50.5L paid"),
            ("Superannuation","","Superannuation",1,        22242,    22242,    "AUD","Australia","Employer + personal contributions"),
        ]
        for inv in investments_seed:
            c.execute("INSERT OR IGNORE INTO investments (name,ticker,inv_type,units,cost_basis,current_value,currency,country,notes) VALUES (?,?,?,?,?,?,?,?,?)", inv)

        # Sample transactions
        now = datetime.date.today()
        ym = now.strftime("%Y-%m")
        expenses = [
            (f"{ym}-01","expense","Rent",        "Monthly Rent",         1733,"AUD",""),
            (f"{ym}-05","expense","Utilities",   "Electricity",           120,"AUD",""),
            (f"{ym}-05","expense","Utilities",   "Gas",                    53,"AUD",""),
            (f"{ym}-05","expense","Utilities",   "Water",                  40,"AUD",""),
            (f"{ym}-05","expense","Internet",    "Internet",               62,"AUD",""),
            (f"{ym}-05","expense","Mobile",      "Mobile Plan",            25,"AUD",""),
            (f"{ym}-05","expense","Insurance",   "Car Insurance",         130,"AUD",""),
            (f"{ym}-05","expense","Car",         "Car Registration",       73,"AUD",""),
            (f"{ym}-05","expense","Health",      "Gym Membership",        223,"AUD",""),
            (f"{ym}-05","expense","Donation",    "Smith Family Donation",  60,"AUD",""),
            (f"{ym}-05","expense","Subscriptions","Subscriptions",         36,"AUD",""),
        ]
        for e in expenses:
            c.execute("INSERT INTO transactions (date,type,category,description,amount,currency,notes) VALUES (?,?,?,?,?,?,?)",
                      (e[0],e[1],e[2],e[3],e[4],e[5],e[6]))

def seed_liabilities():
    with get_db() as c:
        count = c.execute("SELECT COUNT(*) FROM liabilities").fetchone()[0]
        if count > 0:
            return
        c.execute(
            "INSERT INTO liabilities (name,liab_type,amount,currency,country,interest_rate,notes) VALUES (?,?,?,?,?,?,?)",
            ("Family Loan", "Family Loan", 18385, "AUD", "Australia", 0, "Informal loan, no fixed interest/repayment schedule")
        )

def seed_categories():
    with get_db() as c:
        count = c.execute("SELECT COUNT(*) FROM categories").fetchone()[0]
        if count > 0:
            return
        defaults = [
            # income
            ("Salary",       "income",  "#00c9a0", "💼"),
            ("Business Income","income","#00c9a0", "🧾"),
            ("Freelance",    "income",  "#00c9a0", "💻"),
            ("Rental Income","income",  "#00c9a0", "🏠"),
            ("Dividends",    "income",  "#00c9a0", "📈"),
            ("Interest",     "income",  "#00c9a0", "🏦"),
            ("Gift",         "income",  "#00c9a0", "🎁"),
            ("Other Income", "income",  "#00c9a0", "💰"),
            # expense
            ("Rent",         "expense", "#f56565", "🏠"),
            ("Utilities",    "expense", "#f56565", "💡"),
            ("Groceries",    "expense", "#f56565", "🛒"),
            ("Food",         "expense", "#f56565", "🍽️"),
            ("Transport",    "expense", "#f56565", "🚗"),
            ("Internet",     "expense", "#f56565", "🌐"),
            ("Mobile",       "expense", "#f56565", "📱"),
            ("Insurance",    "expense", "#f56565", "🛡️"),
            ("Health",       "expense", "#f56565", "💪"),
            ("Gym",          "expense", "#f56565", "🏋️"),
            ("Subscriptions","expense", "#f56565", "🔔"),
            ("Donation",     "expense", "#f56565", "❤️"),
            ("Car",          "expense", "#f56565", "🚗"),
            ("Shopping",     "expense", "#f56565", "🛍️"),
            ("Travel",       "expense", "#f56565", "✈️"),
            ("Entertainment","expense", "#f56565", "🎬"),
            ("Education",    "expense", "#f56565", "📚"),
            ("Tax",          "expense", "#f56565", "📋"),
            ("Other Expense","expense", "#f56565", "💸"),
            ("Property Maintenance","expense","#e65100","🔧"),
            ("Property Repair","expense","#e65100","🛠️"),
            # investment
            ("US ETF",       "investment", "#7c75f5", "🇺🇸"),
            ("AU ETF",       "investment", "#7c75f5", "🇦🇺"),
            ("Property",     "investment", "#7c75f5", "🏢"),
            ("Crypto",       "investment", "#7c75f5", "₿"),
            ("Stocks",       "investment", "#7c75f5", "📊"),
            ("Fixed Deposit","investment", "#7c75f5", "🏦"),
            ("Savings",      "investment", "#7c75f5", "💰"),
            ("Other Invest", "investment", "#7c75f5", "📈"),
        ]
        for row in defaults:
            c.execute("INSERT OR IGNORE INTO categories (name, type, color, icon) VALUES (?,?,?,?)", row)

def seed_users():
    with get_db() as c:
        if c.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
            c.execute("INSERT INTO users (id,name,emoji,color) VALUES (1,'Pavel','👨','#1565c0')")

def migrate_seed_auth():
    """Idempotent: gives every user a login (username + hashed password) without ever
    touching credentials that already exist. User id=1 always becomes the Superadmin.
    Plaintext defaults only ever exist transiently inside this function, hashed
    immediately via generate_password_hash — never logged, persisted, or kept in memory."""
    with get_db() as c:
        rows = c.execute("SELECT id, name, username, password_hash FROM users ORDER BY id").fetchall()
        existing_usernames = {r["username"] for r in rows if r["username"]}
        for r in rows:
            if r["password_hash"]:
                continue
            if r["id"] == 1:
                uname = "pavelblank"
                n = 2
                while uname in existing_usernames:
                    uname = f"pavelblank{n}"; n += 1
                c.execute("UPDATE users SET username=?, password_hash=?, is_admin=1 WHERE id=1",
                          (uname, generate_password_hash("Pakpak@123")))
            else:
                base = re.sub(r'[^a-z0-9]+', '', (r["name"] or "user").lower()) or "user"
                uname = base; n = 2
                while uname in existing_usernames:
                    uname = f"{base}{n}"; n += 1
                c.execute("UPDATE users SET username=?, password_hash=? WHERE id=?",
                          (uname, generate_password_hash("admin123"), r["id"]))
            existing_usernames.add(uname)

def _get_or_create_demo_user():
    """Ensures the dedicated, public Demo account (username/password: demo/demo) exists and
    returns its user id. Idempotent - safe to call on every startup. Runs after
    migrate_seed_auth() and inserts its password_hash directly, so that function's
    no-password-yet backfill loop never touches this row."""
    with get_db() as c:
        row = c.execute("SELECT id FROM users WHERE LOWER(username)=?", (DEMO_USERNAME,)).fetchone()
        if row:
            return row["id"]
        cur = c.execute(
            "INSERT INTO users (name,emoji,color,username,password_hash,is_admin) VALUES (?,?,?,?,?,0)",
            ("Demo", "🎭", "#a78bfa", DEMO_USERNAME, generate_password_hash(DEMO_PASSWORD))
        )
        return cur.lastrowid

seed_users()
migrate_seed_auth()
_DEMO_UID = _get_or_create_demo_user()
seed_data()
seed_categories()
seed_liabilities()
migrate_fix_known_data()

# ─── API: USERS ──────────────────────────────────────────────────────────────

@app.route("/api/users", methods=["GET"])
def api_get_users():
    caller = session.get('user_id')
    with get_db() as c:
        rows = c.execute("SELECT id, name, emoji, color, tfn, username, is_admin FROM users ORDER BY id").fetchall()
    data = [row_to_dict(r) for r in rows]
    # Blank every TFN unless the caller is a logged-in admin (sees all) or a logged-in
    # non-admin viewing their own record. Previously this only ran "if caller and not
    # admin", which left TFNs fully exposed to anonymous callers - since /api/* stays
    # open for the local pavel-ai-stack MCP integration, that meant anyone who could
    # reach this endpoint (no session needed) got every household member's real TFN.
    # The MCP's own tools (fin_summary/fin_accounts/fin_transactions/etc.) never read
    # this field, so blanking it for anonymous callers changes nothing for that integration.
    if not _is_admin():
        for u in data:
            if u["id"] != caller:
                u["tfn"] = ""
    return jsonify({"ok":True,"data":data})

@app.route("/api/users", methods=["POST"])
def api_add_user():
    if not _is_admin():
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    d = request.json
    if not d.get("name"): return jsonify({"ok":False,"error":"Name required"})
    with get_db() as c:
        existing = {r["username"] for r in c.execute("SELECT username FROM users WHERE username IS NOT NULL").fetchall()}
        base = re.sub(r'[^a-z0-9]+', '', d["name"].strip().lower()) or "user"
        uname = base; n = 2
        while uname in existing:
            uname = f"{base}{n}"; n += 1
        cur = c.execute("INSERT INTO users (name,emoji,color,tfn,username,password_hash,is_admin) VALUES (?,?,?,?,?,?,0)",
            (d["name"].strip(), d.get("emoji","👤"), d.get("color","#1565c0"), d.get("tfn","").strip(),
             uname, generate_password_hash("admin123")))
        row = c.execute("SELECT id, name, emoji, color, tfn, username, is_admin FROM users WHERE id=?", (cur.lastrowid,)).fetchone()
    return jsonify({"ok":True,"data":row_to_dict(row)})

@app.route("/api/users/<int:uid>", methods=["PUT"])
def api_update_user(uid):
    if not _is_admin():
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    d = request.json
    with get_db() as c:
        if "tfn" in d:
            c.execute("UPDATE users SET name=?,emoji=?,color=?,tfn=? WHERE id=?",
                (d["name"], d.get("emoji","👤"), d.get("color","#1565c0"), d.get("tfn","").strip(), uid))
        else:
            c.execute("UPDATE users SET name=?,emoji=?,color=? WHERE id=?",
                (d["name"], d.get("emoji","👤"), d.get("color","#1565c0"), uid))
        row = c.execute("SELECT id, name, emoji, color, tfn, username, is_admin FROM users WHERE id=?", (uid,)).fetchone()
    return jsonify({"ok":True,"data":row_to_dict(row) if row else None})

@app.route("/api/users/<int:uid>", methods=["DELETE"])
def api_delete_user(uid):
    if not _is_admin():
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    if uid == 1: return jsonify({"ok":False,"error":"Cannot delete primary user"})
    with get_db() as c:
        c.execute("UPDATE transactions SET user_id=1 WHERE user_id=?", (uid,))
        c.execute("UPDATE accounts SET user_id=1 WHERE user_id=?", (uid,))
        # Same reassign-on-delete pattern, extended to investments/liabilities/businesses
        # now that they also carry per-user ownership.
        c.execute("UPDATE investments SET user_id=1 WHERE user_id=?", (uid,))
        c.execute("UPDATE liabilities SET user_id=1 WHERE user_id=?", (uid,))
        c.execute("UPDATE businesses SET user_id=1 WHERE user_id=?", (uid,))
        c.execute("DELETE FROM users WHERE id=?", (uid,))
    return jsonify({"ok":True})

@app.route("/api/set-user", methods=["POST"])
def api_set_user():
    if not _is_admin():
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    uid = request.json.get("uid")
    if uid is None: session.pop("view_uid", None)
    else: session["view_uid"] = int(uid)
    return jsonify({"ok":True})

# ─── API: RECURRING TEMPLATES ────────────────────────────────────────────────

def _next_date(current_date, frequency):
    """Calculate next occurrence date based on frequency."""
    d = datetime.date.fromisoformat(current_date)
    if frequency == 'weekly':
        return (d + datetime.timedelta(weeks=1)).isoformat()
    elif frequency == 'monthly':
        month = d.month + 1
        year = d.year
        if month > 12:
            month = 1
            year += 1
        day = min(d.day, 28)
        return datetime.date(year, month, day).isoformat()
    elif frequency == 'yearly':
        return datetime.date(d.year + 1, d.month, min(d.day, 28)).isoformat()
    return current_date

def process_recurring():
    """Create transactions from due recurring templates. Returns count of created transactions."""
    today = datetime.date.today().isoformat()
    created = 0
    with get_db() as c:
        templates = c.execute(
            "SELECT * FROM recurring_templates WHERE active=1 AND next_date<=?"
            " ORDER BY next_date ASC", (today,)
        ).fetchall()
        for tpl in templates:
            tpl = dict(tpl)
            # Check if end_date has passed
            if tpl['end_date'] and tpl['end_date'] < today:
                c.execute("UPDATE recurring_templates SET active=0 WHERE id=?", (tpl['id'],))
                continue
            # Create the transaction linked to the recurring template
            c.execute(
                "INSERT INTO transactions (date,type,category,description,amount,currency,account,notes,user_id,recurring_id) VALUES (?,?,?,?,?,?,?,?,?,?)",
                (tpl['next_date'], tpl['type'], tpl['category'], tpl['description'],
                 tpl['amount'], tpl['currency'], tpl['account'], tpl['notes'], tpl['user_id'], tpl['id'])
            )
            # Calculate next_date
            new_next = _next_date(tpl['next_date'], tpl['frequency'])
            # If end_date set and next exceeds it, deactivate
            if tpl['end_date'] and new_next > tpl['end_date']:
                c.execute("UPDATE recurring_templates SET active=0 WHERE id=?", (tpl['id'],))
            else:
                c.execute("UPDATE recurring_templates SET next_date=? WHERE id=?", (new_next, tpl['id']))
            created += 1
    return created

@app.route("/api/recurring", methods=["GET"])
def api_get_recurring():
    uid = _uid()
    with get_db() as c:
        wh, p = _uf(uid)
        rows = c.execute(f"SELECT * FROM recurring_templates{wh} ORDER BY active DESC, next_date ASC", p).fetchall()
    return jsonify({"ok": True, "data": [row_to_dict(r) for r in rows]})

@app.route("/api/recurring", methods=["POST"])
def api_add_recurring():
    d = request.json
    required = ["type", "category", "amount", "frequency", "start_date"]
    if not all(d.get(k) for k in required):
        return jsonify({"ok": False, "error": "Missing required fields"})
    if d["frequency"] not in ("weekly", "monthly", "yearly"):
        return jsonify({"ok": False, "error": "Frequency must be weekly, monthly, or yearly"})
    uid_in = d.get("user_id") if _can_set_user_id() else None
    uid = int(uid_in) if uid_in not in (None, "") else (_uid() or 1)
    # Auto-calculate next_date from start_date
    start = d["start_date"]
    today = datetime.date.today().isoformat()
    next_date = start if start >= today else today
    with get_db() as c:
        cur = c.execute(
            "INSERT INTO recurring_templates (type,category,description,amount,currency,account,notes,user_id,frequency,start_date,end_date,next_date) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            (d["type"], d["category"], d.get("description", ""), float(d["amount"]),
             d.get("currency", "AUD"), d.get("account", ""), d.get("notes", ""),
             uid, d["frequency"], start, d.get("end_date"), next_date)
        )
        row = c.execute("SELECT * FROM recurring_templates WHERE id=?", (cur.lastrowid,)).fetchone()
    return jsonify({"ok": True, "data": row_to_dict(row)})

@app.route("/api/recurring/<int:rid>", methods=["PUT"])
def api_update_recurring(rid):
    d = request.json
    with get_db() as c:
        existing = c.execute("SELECT user_id FROM recurring_templates WHERE id=?", (rid,)).fetchone()
        if existing and _forbidden_if_not_owner(existing["user_id"]):
            return jsonify({"ok": False, "error": "Forbidden"}), 403
        c.execute(
            "UPDATE recurring_templates SET type=?,category=?,description=?,amount=?,currency=?,account=?,notes=?,frequency=?,start_date=?,end_date=?,active=? WHERE id=?",
            (d["type"], d["category"], d.get("description", ""), float(d["amount"]),
             d.get("currency", "AUD"), d.get("account", ""), d.get("notes", ""),
             d["frequency"], d["start_date"], d.get("end_date"), int(d.get("active", 1)), rid)
        )
        row = c.execute("SELECT * FROM recurring_templates WHERE id=?", (rid,)).fetchone()
    return jsonify({"ok": True, "data": row_to_dict(row) if row else None})

@app.route("/api/recurring/<int:rid>", methods=["DELETE"])
def api_delete_recurring(rid):
    with get_db() as c:
        row = c.execute("SELECT user_id FROM recurring_templates WHERE id=?", (rid,)).fetchone()
        if row and _forbidden_if_not_owner(row["user_id"]):
            return jsonify({"ok": False, "error": "Forbidden"}), 403
        # Remove all transactions generated from this recurring template
        c.execute("DELETE FROM transactions WHERE recurring_id=?", (rid,))
        c.execute("DELETE FROM recurring_templates WHERE id=?", (rid,))
    return jsonify({"ok": True})

@app.route("/api/recurring/process", methods=["POST"])
def api_process_recurring():
    # Global, no uid filter — would create real transactions for the real household.
    # The public Demo account has no legitimate reason to call this.
    if _uid() == _DEMO_UID:
        return jsonify({"ok": False, "error": "Not available in demo mode"}), 403
    count = process_recurring()
    return jsonify({"ok": True, "created": count})

# ─── API: CATEGORIES ─────────────────────────────────────────────────────────

@app.route("/api/categories", methods=["GET"])
def api_get_categories():
    type_f = request.args.get("type", "")
    with get_db() as c:
        if type_f:
            rows = c.execute("SELECT * FROM categories WHERE type=? ORDER BY name", (type_f,)).fetchall()
        else:
            rows = c.execute("SELECT * FROM categories ORDER BY type, name").fetchall()
    return jsonify({"ok": True, "data": [row_to_dict(r) for r in rows]})

@app.route("/api/categories", methods=["POST"])
def api_add_category():
    # categories has zero per-user scoping — it's one shared list for the whole household.
    # Not reachable from any page Demo can open (only Settings has this UI, and Settings is
    # blocked for Demo), but guarded directly too since it has no auth otherwise.
    if _uid() == _DEMO_UID:
        return jsonify({"ok": False, "error": "Not available in demo mode"}), 403
    d = request.json
    name = (d.get("name") or "").strip()
    cat_type = (d.get("type") or "expense").strip()
    if not name:
        return jsonify({"ok": False, "error": "Name is required"})
    with get_db() as c:
        try:
            cur = c.execute(
                "INSERT INTO categories (name, type, color, icon) VALUES (?,?,?,?)",
                (name, cat_type, d.get("color","#7c75f5"), d.get("icon",""))
            )
            row = c.execute("SELECT * FROM categories WHERE id=?", (cur.lastrowid,)).fetchone()
            return jsonify({"ok": True, "data": row_to_dict(row)})
        except sqlite3.IntegrityError:
            row = c.execute("SELECT * FROM categories WHERE name=? AND type=?", (name, cat_type)).fetchone()
            return jsonify({"ok": True, "data": row_to_dict(row), "existing": True})

@app.route("/api/categories/<int:cid>", methods=["DELETE"])
def api_delete_category(cid):
    # Same shared, unscoped table as above — would let Demo delete a real category still
    # used by the real household's real transactions.
    if _uid() == _DEMO_UID:
        return jsonify({"ok": False, "error": "Not available in demo mode"}), 403
    with get_db() as c:
        c.execute("DELETE FROM categories WHERE id=?", (cid,))
    return jsonify({"ok": True})

@app.route("/api/categories/ensure", methods=["POST"])
def api_ensure_category():
    """Auto-add a category if it doesn't exist (called from transaction form)."""
    d = request.json
    name = (d.get("name") or "").strip()
    cat_type = (d.get("type") or "expense").strip()
    if not name:
        return jsonify({"ok": False, "error": "Name required"})
    with get_db() as c:
        row = c.execute("SELECT * FROM categories WHERE name=? AND type=?", (name, cat_type)).fetchone()
        if not row and _uid() == _DEMO_UID:
            # categories is shared/unscoped — a Demo visitor typing a brand-new category
            # name in the transaction form must never permanently add it to the real
            # household's list. Hand back a synthetic, never-persisted category instead so
            # Demo's own add-transaction flow still works.
            return jsonify({"ok": True, "data": {"id": 0, "name": name, "type": cat_type, "color": "#a78bfa", "icon": ""}})
        if not row:
            c.execute(
                "INSERT OR IGNORE INTO categories (name, type, color, icon) VALUES (?,?,?,?)",
                (name, cat_type, "#a78bfa", "")
            )
            row = c.execute("SELECT * FROM categories WHERE name=? AND type=?", (name, cat_type)).fetchone()
    return jsonify({"ok": True, "data": row_to_dict(row)})

# ── Notes ──────────────────────────────────────────────────────────────────────

@app.route("/notes")
def notes_page():
    # Notes are entirely shared/global with no per-user scoping - never let the public
    # Demo account anywhere near the household's real notes.
    if _uid() == _DEMO_UID:
        return redirect('/')
    return render_template("notes.html")

def _require_session_json():
    """Shared guard for /api/ endpoints that expose real, sensitive household data with no
    per-user scoping (notes, tax notes, audit log, business profile, businesses). These stay
    under /api/ for path consistency, but unlike the finance CRUD endpoints, they were never
    meant to be reachable without a login - the local pavel-ai-stack MCP integration has no
    tool that reads or writes any of them, so requiring a session here doesn't affect it.
    Also blocks the public Demo account, since none of this data belongs to it either.
    Returns an error response to return immediately, or None if the caller may proceed."""
    uid = session.get('user_id')
    if not uid:
        return jsonify({"ok": False, "error": "Login required"}), 401
    if uid == _DEMO_UID:
        return jsonify({"ok": False, "error": "Not available in demo mode"}), 403
    return None

@app.route("/api/notes", methods=["GET"])
def api_get_notes():
    _err = _require_session_json()
    if _err: return _err
    with get_db() as c:
        rows = c.execute("SELECT * FROM notes ORDER BY pinned DESC, updated DESC").fetchall()
    return jsonify({"ok": True, "data": [row_to_dict(r) for r in rows]})

@app.route("/api/notes", methods=["POST"])
def api_add_note():
    _err = _require_session_json()
    if _err: return _err
    with get_db() as c:
        count = c.execute("SELECT COUNT(*) as n FROM notes").fetchone()["n"]
    if count >= 3:
        return jsonify({"ok": False, "error": "Maximum 3 notes allowed"})
    d = request.json
    title = (d.get("title") or "").strip()
    content = (d.get("content") or "").strip()
    color = (d.get("color") or "#a78bfa").strip()
    with get_db() as c:
        cur = c.execute(
            "INSERT INTO notes (title, content, color) VALUES (?,?,?)",
            (title, content, color)
        )
        row = c.execute("SELECT * FROM notes WHERE id=?", (cur.lastrowid,)).fetchone()
    return jsonify({"ok": True, "data": row_to_dict(row)})

@app.route("/api/notes/<int:nid>", methods=["PUT"])
def api_update_note(nid):
    _err = _require_session_json()
    if _err: return _err
    d = request.json
    with get_db() as c:
        existing = c.execute("SELECT * FROM notes WHERE id=?", (nid,)).fetchone()
        if not existing:
            return jsonify({"ok": False, "error": "Note not found"})
        title = (d.get("title") if d.get("title") is not None else existing["title"]).strip()
        content = (d.get("content") if d.get("content") is not None else existing["content"]).strip()
        color = (d.get("color") if d.get("color") is not None else existing["color"]).strip()
        pinned = d.get("pinned") if d.get("pinned") is not None else existing["pinned"]
        c.execute(
            "UPDATE notes SET title=?, content=?, color=?, pinned=?, updated=datetime('now') WHERE id=?",
            (title, content, color, pinned, nid)
        )
        row = c.execute("SELECT * FROM notes WHERE id=?", (nid,)).fetchone()
    return jsonify({"ok": True, "data": row_to_dict(row)})

@app.route("/api/notes/<int:nid>", methods=["DELETE"])
def api_delete_note(nid):
    _err = _require_session_json()
    if _err: return _err
    with get_db() as c:
        c.execute("DELETE FROM notes WHERE id=?", (nid,))
    return jsonify({"ok": True})

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=8082, debug=False)
